from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import *
from django.contrib.auth.models import User
from django.db.models import Max, Q, Sum, F, Count, Avg, Subquery, OuterRef
import pandas as pd
from decimal import Decimal
import openpyxl
from datetime import datetime, timedelta
from django.utils import timezone
from django.db import transaction
from django.db.models.functions import TruncMonth, TruncWeek, TruncDate
import json
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from io import BytesIO
from PIL import Image
from django.core.files.base import ContentFile
from django.http import JsonResponse




def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            request.session['username'] = username  # Set session data
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    return render(request, 'core/login.html')





@login_required(login_url='login')
def dashboard(request):
    # Redirect based on user role
    if request.user.profile.role == 'Doctor':
        return redirect('doctor_dashboard')
    elif request.user.profile.role == 'Nurse':
        return redirect('nurse_dashboard')
    elif request.user.profile.role == 'LabTech':
        return redirect('labtech_dashboard')
    elif request.user.profile.role == 'Pharmacist':
        return redirect('pharmacy_dashboard')
    elif request.user.profile.role == 'PharmacyHead':
        return redirect('pharmacy_dashboard')
    elif request.user.profile.role == 'Receptionist':
        return redirect('receptionist_dashboard')
    elif request.user.profile.role == 'Cashier':
        return redirect('cashier_dashboard')
    elif request.user.profile.role == 'Accountant':
        return redirect('cashier_dashboard')

    # Fetch the metrics for the dashboard
    today = timezone.now().date()

    # New bookings for today
    new_bookings = Appointment.objects.filter(appointment_date__date=today).count()

    # Total active customers (patients)
    total_customers = Patient.objects.filter(is_active=True).count()

    # New projects (admissions) for today
    new_projects = Admission.objects.filter(admission_date__date=today).count()

    # Total revenue for today
    total_revenue = Payment.objects.filter(payment_date__date=today).aggregate(total=Sum('amount'))['total'] or 0

    # Monthly revenue data for the chart (use TruncMonth)
    start_of_year = datetime(today.year, 1, 1)
    monthly_revenue = Payment.objects.filter(payment_date__gte=start_of_year) \
        .annotate(month=TruncMonth('payment_date')) \
        .values('month') \
        .annotate(total=Sum('amount')) \
        .order_by('month')

    revenue_chart_data = {
        'months': [month['month'].strftime('%B') for month in monthly_revenue],
        'totals': [month['total'] for month in monthly_revenue]
    }

    # Weekly customer growth chart data (last 6 weeks)
    six_weeks_ago = today - timedelta(weeks=6)
    weekly_customers = Patient.objects.filter(created_at__gte=six_weeks_ago) \
        .annotate(week=TruncWeek('created_at')) \
        .values('week') \
        .annotate(count=Count('id')) \
        .order_by('week')

    customer_growth_chart_data = {
        'weeks': [week['week'].strftime('%U') for week in weekly_customers],  # Use %U to represent week number
        'counts': [week['count'] for week in weekly_customers]
    }

    # Ongoing tasks or appointments table data
    ongoing_appointments = Appointment.objects.filter(status='Scheduled').order_by('appointment_date')[:10]

    # Recent payments for the revenue table
    recent_payments = Payment.objects.order_by('-payment_date')[:10]

    context = {
        'new_bookings': new_bookings,
        'total_customers': total_customers,
        'new_projects': new_projects,
        'total_revenue': total_revenue,
        'revenue_chart_data': revenue_chart_data,
        'customer_growth_chart_data': customer_growth_chart_data,
        'ongoing_appointments': ongoing_appointments,
        'recent_payments': recent_payments,
    }

    return render(request, 'core/dashboard.html', context)




@login_required(login_url='login')
def user_logout(request):
    logout(request)
    request.session.flush()  # Clear all session data
    return redirect('login')

def register(request):
    if request.user.profile.role == 'Admin':
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            email = request.POST.get('email')
            role = request.POST.get('role')
            contact_number = request.POST.get('contact_number')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            address = request.POST.get('address')
            profile_picture = request.FILES.get('profile_picture')

            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already taken.')
            else:
                user = User.objects.create_user(username=username, password=password, email=email, first_name=first_name, last_name=last_name)
                profile = Profile(user=user, role=role, contact_number=contact_number, address=address, profile_picture=profile_picture)
                profile.save()
                messages.success(request, 'Registration successful.')
                return redirect('manage_user_profiles')  # Redirect to the dashboard after registration

        return render(request, 'core/register.html')
    return redirect('dashboard')

@login_required(login_url='login')
def update_profile(request, profile_id=None):
    
    if profile_id:
        profile = Profile.objects.get(id=profile_id)
        user = profile.user
    
    else:
        
        user = request.user
        profile = user.profile

    if request.method == 'POST':
        user.email = request.POST.get('email')
        profile.role = request.POST.get('role')
        profile.contact_number = request.POST.get('contact_number')
        profile.address = request.POST.get('address')

        if request.FILES.get('profile_picture'):
            profile.profile_picture = request.FILES.get('profile_picture')

        user.save()
        profile.save()
        messages.success(request, 'Profile updated successfully.')
        if profile_id:
            return redirect('manage_user_profiles')  # Redirect to the profile page after update
        else:
            return redirect('update_profile')  # Redirect to the profile page after update

    return render(request, 'core/update_profile.html', {'user': user, 'profile': profile})

# ===================================Receptionist=======================



@login_required(login_url='login')
def receptionist_dashboard(request):
    # Check if the user is allowed to view this page
    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Receptionist':
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        patients = Patient.objects.filter(created_at__date__range=[start_date, end_date]).order_by('-created_at')

        return render(request, 'core/receptionist_dashboard.html', {
            'patients': patients,
            'start_date': start_date,
            'end_date': end_date,
        })

    return redirect('dashboard')



@login_required(login_url='login')
def create_family(request):
    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Receptionist':
        if request.method == 'POST':
            name = request.POST.get('name')
            file_number = request.POST.get('file_number')
            address = request.POST.get('address')
            phone_number = request.POST.get('phone_number')

            if name and file_number:
                # Create and save the Family object
                family = Family.objects.create(
                    name=name,
                    file_number=file_number,
                    address=address,
                    phone_number=phone_number
                )
                
                # Create a wallet for the family
                Wallet.objects.create(family=family, balance=0.00)

                return redirect('family_list')  # Redirect to a success page or family list view

        return render(request, 'core/create_family.html')
    
    return redirect('dashboard')



@login_required(login_url='login')
def family_list(request):
    # Check if the user is allowed to view this page
    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Receptionist':
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        families = Family.objects.filter(created_at__date__range=[start_date, end_date]).order_by('-created_at')

        return render(request, 'core/family_list.html', {
            'families': families,
            'start_date': start_date,
            'end_date': end_date,
        })

    return redirect('dashboard')


@login_required(login_url='login')
def view_family_members(request, family_id):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        patients = Patient.objects.filter(family=Family.objects.get(id=family_id)).all()

        return render(request, 'core/family_members.html', context={'patients': patients})
    return redirect('dashboard')
    



@login_required(login_url='login')
def edit_family(request, family_id):
    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Receptionist':
        family = get_object_or_404(Family, id=family_id)

        if request.method == 'POST':
            family.name = request.POST.get('name')
            family.file_number = request.POST.get('file_number')
            family.address = request.POST.get('address')
            family.phone_number = request.POST.get('phone_number')
            family.save()
            return redirect('family_list')

        return render(request, 'core/edit_family.html', {'family': family})
    return redirect('dashboard')



@login_required(login_url='login')
def delete_family(request, family_id):
    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Receptionist':
        family = get_object_or_404(Family, id=family_id)
        family.delete()
        return redirect('family_list')
    return redirect('dashboard')
    

@login_required(login_url='login')
def create_patient(request):
    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Receptionist':
        if request.method == 'POST':
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            age = request.POST.get('age')
            gender = request.POST.get('gender')
            contact_number = request.POST.get('contact_number')
            address = request.POST.get('address')
            next_of_kin_name = request.POST.get('next_of_kin_name')
            next_of_kin_relationship = request.POST.get('next_of_kin_relationship')
            next_of_kin_contact = request.POST.get('next_of_kin_contact')
            patient_file_number = request.POST.get('patient_file_number')
            family_id = request.POST.get('family')
            file_type = request.POST.get('file_type')
            hmo_id = request.POST.get('hmo')
            company_id = request.POST.get('company')
            profile_picture = request.FILES.get('profile_picture')

            # Fetch Family if selected
            if family_id:
                family = Family.objects.get(id=family_id)
            else:
                family = None

            hmo = HMO.objects.filter(id=hmo_id).first() if hmo_id else None
            company = Company.objects.filter(id=company_id).first() if company_id else None

            # Automate patient file number if not provided
            if not patient_file_number:
                all_patients = Patient.objects.all()
                all_number = [int(i.patient_file_number.replace('AHZ', '').replace('OLD', '')) for i in all_patients]
                last_number = max(all_number) if all_number else 1000  # Set default to AHZ1001
                patient_file_number = f'AHZ{last_number + 1}'

            # Create the patient object
            patient = Patient.objects.create(
                first_name=first_name,
                last_name=last_name,
                age=age,
                gender=gender,
                created_by=request.user.profile,
                contact_number=contact_number,
                address=address,
                patient_file_number=patient_file_number,
                file_type=file_type,
                hmo=hmo,
                company=company,
                profile_picture=profile_picture,
                next_of_kin_name=next_of_kin_name,
                next_of_kin_relationship=next_of_kin_relationship,
                next_of_kin_contact=next_of_kin_contact,
                family=family,
            )

            # If the patient is not associated with a family, create a wallet for the patient
            if not family:
                Wallet.objects.create(patient=patient, balance=0.00)

            # Create an invoice for the new patient registration
            try:
                service = Service.objects.get(name='New Patient Registration')
                # Create the invoice
                if family:
                    invoice = Invoice.objects.create(
                        family=family,
                        total_amount=service.cost,
                        is_paid=False
                    )
                else:
                    invoice = Invoice.objects.create(
                        patient=patient,
                        total_amount=service.cost,
                        is_paid=False
                    )

                # Create an invoice item linked to the service
                InvoiceItem.objects.create(
                    invoice=invoice,
                    service=service,
                    quantity=1,  # Typically one registration fee per patient
                    unit_price=service.cost,
                    total_price=service.cost
                )
            except Service.DoesNotExist:
                messages.error(request, 'Error creating invoice for new patient. Please check the services setup.')

            messages.success(request, 'Patient created successfully.')
            return redirect('receptionist_dashboard')
        
        context = {
            'families': Family.objects.all(),
            'hmos': HMO.objects.all().order_by('name'),
            'companies': Company.objects.all().order_by('name'),
        }

        return render(request, 'core/create_patient.html', context)
    
    return redirect('dashboard')


@login_required(login_url='login')
def update_patient(request, patient_id):
    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Receptionist':
        patient = Patient.objects.get(id=patient_id)

        if request.method == 'POST':
            patient.first_name = request.POST.get('first_name')
            patient.last_name = request.POST.get('last_name')
            patient.age = request.POST.get('age')
            patient.gender = request.POST.get('gender')
            patient.contact_number = request.POST.get('contact_number')
            patient.address = request.POST.get('address')
            patient.next_of_kin_name = request.POST.get('next_of_kin_name')
            patient.next_of_kin_relationship = request.POST.get('next_of_kin_relationship')
            patient.next_of_kin_contact = request.POST.get('next_of_kin_contact')
            patient.patient_file_number = request.POST.get('patient_file_number')
            patient.file_type = request.POST.get('file_type')

            hmo_id = request.POST.get('hmo')
            patient.hmo = HMO.objects.filter(id=hmo_id).first() if hmo_id else None

            company_id = request.POST.get('company')
            patient.company = Company.objects.filter(id=company_id).first() if company_id else None

            if request.FILES.get('profile_picture'):
                patient.profile_picture = request.FILES.get('profile_picture')

            try:
                if request.POST.get('family') == "None":
                    patient.family = None
                else:
                    patient.family = Family.objects.get(id=request.POST.get('family'))
            except:
                pass


            patient.save()
            messages.success(request, 'Patient record updated successfully.')
            return redirect('receptionist_dashboard')

        return render(request, 'core/update_patient.html', {
            'patient': patient,
            'families': Family.objects.all(),
            'hmos': HMO.objects.all().order_by('name'),
            'companies': Company.objects.all().order_by('name'),
        })

    return redirect('dashboard')


@login_required(login_url='login')
def delete_patient(request, patient_id):
    Patient.objects.get(id=patient_id).delete()
    return redirect('receptionist_dashboard')


# ========================HMO Management============================

@login_required(login_url='login')
def hmo_list(request):
    if request.user.profile.role not in ['Admin', 'Receptionist']:
        return redirect('dashboard')

    hmos = HMO.objects.all().order_by('name')
    return render(request, 'core/hmo_list.html', {'hmos': hmos})


@login_required(login_url='login')
def hmo_form(request, hmo_id=None):
    if request.user.profile.role not in ['Admin', 'Receptionist']:
        return redirect('dashboard')

    hmo = get_object_or_404(HMO, id=hmo_id) if hmo_id else None

    if request.method == 'POST':
        name = request.POST.get('name')
        if hmo:
            hmo.name = name
            hmo.save()
            messages.success(request, 'HMO updated successfully.')
        else:
            HMO.objects.create(name=name)
            messages.success(request, 'HMO created successfully.')
        return redirect('hmo_list')

    return render(request, 'core/hmo_form.html', {'hmo': hmo})


@login_required(login_url='login')
def hmo_delete(request, hmo_id):
    if request.user.profile.role not in ['Admin', 'Receptionist']:
        return redirect('dashboard')

    get_object_or_404(HMO, id=hmo_id).delete()
    messages.success(request, 'HMO deleted successfully.')
    return redirect('hmo_list')


# ========================Company Management============================

@login_required(login_url='login')
def company_list(request):
    if request.user.profile.role not in ['Admin', 'Receptionist']:
        return redirect('dashboard')

    companies = Company.objects.all().order_by('name')
    return render(request, 'core/company_list.html', {'companies': companies})


@login_required(login_url='login')
def company_form(request, company_id=None):
    if request.user.profile.role not in ['Admin', 'Receptionist']:
        return redirect('dashboard')

    company = get_object_or_404(Company, id=company_id) if company_id else None

    if request.method == 'POST':
        name = request.POST.get('name')
        if company:
            company.name = name
            company.save()
            messages.success(request, 'Company updated successfully.')
        else:
            Company.objects.create(name=name)
            messages.success(request, 'Company created successfully.')
        return redirect('company_list')

    return render(request, 'core/company_form.html', {'company': company})


@login_required(login_url='login')
def company_delete(request, company_id):
    if request.user.profile.role not in ['Admin', 'Receptionist']:
        return redirect('dashboard')

    get_object_or_404(Company, id=company_id).delete()
    messages.success(request, 'Company deleted successfully.')
    return redirect('company_list')


# ========================Creating Appointment============================

@login_required(login_url='login')
def create_appointment(request):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        if request.method == 'POST':
            patient_id = request.POST.get('patient')
            doctor_id = request.POST.get('doctor')
            appointment_date = request.POST.get('appointment_date')
            reason = request.POST.get('reason')
            reason_category = request.POST.get('reason_category')
            status = request.POST.get('status', 'Scheduled')

            # Fetch the related objects
            patient = Patient.objects.get(id=patient_id)
            doctor = Doctor.objects.get(id=doctor_id)

            # Create and save the appointment
            appointment = Appointment.objects.create(
                patient=patient,
                doctor=doctor,
                created_by=request.user.profile,
                appointment_date=appointment_date,
                reason=reason,
                reason_category=reason_category,
                status=status
            )

            # Fetch the service related to the reason_category
            try:
                service = Service.objects.get(name=reason_category)
                # Create the invoice
                if patient.family != None:
                    invoice = Invoice.objects.create(
                        family=patient.family,
                        total_amount=service.cost,
                        is_paid=False
                    )
                else:
                    invoice = Invoice.objects.create(
                        patient=patient,
                        total_amount=service.cost,
                        is_paid=False
                    )

                # Create an invoice item linked to the service
                InvoiceItem.objects.create(
                    invoice=invoice,
                    service=service,
                    quantity=1,
                    unit_price=service.cost,
                    total_price=service.cost
                )
            except Service.DoesNotExist:
                messages.error(request, f'Service for {reason_category} not found. Please check the services setup.')

            return redirect('appointment_list')  # Redirect to the list of appointments

        patients = Patient.objects.all()
        doctors = Doctor.objects.all()
        return render(request, 'core/create_appointment.html', {'patients': patients, 'doctors': doctors})
    return redirect('dashboard')


@login_required(login_url='login')
def appointment_list(request):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        appointments = Appointment.objects.filter(created_at__date__range=[start_date, end_date]).order_by('-created_at', '-updated_at')

        return render(request, 'core/appointment_list.html', {
            'appointments': appointments,
            'start_date': start_date,
            'end_date': end_date,
        })

    return redirect('dashboard')

@login_required(login_url='login')
def update_appointment(request, appointment_id):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']: 
        appointment = get_object_or_404(Appointment, id=appointment_id)
        if request.method == 'POST':
            # Handle form submission to update the appointment
            appointment.status = request.POST.get('status')
            appointment.save()
            return redirect('appointment_list')
        return render(request, 'core/update_appointment.html', {'appointment': appointment})
    return redirect('dashboard')

@login_required(login_url='login')
def view_appointment(request, appointment_id):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        # Ensure the user is authorized to view the appointment details
        appointment = get_object_or_404(Appointment, id=appointment_id)

        # Get related prescriptions, investigation requests, admissions, doctor notes, diagnoses, and vital signs
        prescriptions = Prescription.objects.filter(appointment=appointment)
        investigation_requests = InvestigationRequest.objects.filter(appointment=appointment)
        admission = Admission.objects.filter(patient=appointment.patient).first()  # Assuming one admission per patient
        doctor_notes = DoctorNote.objects.filter(appointment=appointment).order_by('-created_at')
        vital_signs = VitalSign.objects.filter(appointment=appointment)
        diagnoses = Diagnosis.objects.filter(appointment=appointment).order_by('-created_at')

        # Full patient history: everything from past appointments and admissions,
        # excluding the current appointment's own records (already shown above)
        patient = appointment.patient
        history_diagnoses = Diagnosis.objects.filter(
            Q(appointment__patient=patient) | Q(admission__patient=patient)
        ).exclude(appointment=appointment).select_related('doctor', 'appointment', 'admission').order_by('-created_at')

        history_prescriptions = Prescription.objects.filter(
            Q(appointment__patient=patient) | Q(admission__patient=patient)
        ).exclude(appointment=appointment).select_related('medication', 'appointment', 'admission').order_by('-date_prescribed')

        history_investigation_requests = InvestigationRequest.objects.filter(
            patient=patient
        ).exclude(appointment=appointment).select_related('investigation').order_by('-date_requested')

        history_admissions = Admission.objects.filter(patient=patient).order_by('-admission_date')

        history_doctor_notes = DoctorNote.objects.filter(
            Q(appointment__patient=patient) | Q(admission__patient=patient)
        ).exclude(appointment=appointment).select_related('doctor', 'appointment', 'admission').order_by('-created_at')

        context = {
            'appointment': appointment,
            'prescriptions': prescriptions,
            'investigation_requests': investigation_requests,
            'admission': admission,
            'doctor_notes': doctor_notes,
            'vital_signs': vital_signs,
            'diagnoses': diagnoses,
            'history_diagnoses': history_diagnoses,
            'history_prescriptions': history_prescriptions,
            'history_investigation_requests': history_investigation_requests,
            'history_admissions': history_admissions,
            'history_doctor_notes': history_doctor_notes,
        }

        return render(request, 'core/view_appointment.html', context)

    return redirect('dashboard')



# ======================= Observation Chart ==============================
@login_required(login_url='login')
def create_observation_chart(request, admission_id):

    if request.user.profile.role == 'Nurse' or request.user.profile.role == 'Admin':
        admission = get_object_or_404(Admission, id=admission_id)

        if request.method == 'POST':
            temperature = request.POST.get('temperature')
            pulse = request.POST.get('pulse')
            respiratory_rate = request.POST.get('respiratory_rate')
            blood_pressure = request.POST.get('blood_pressure')
            oxygen_saturation = request.POST.get('oxygen_saturation')
            remark = request.POST.get('remark')

            # Create the ObservationChart instance
            observation_chart = ObservationChart.objects.create(
                admission=admission,
                temperature=temperature,
                pulse=pulse,
                respiratory_rate=respiratory_rate,
                blood_pressure=blood_pressure,
                oxygen_saturation=oxygen_saturation,
                remark=remark,
                recorded_by=request.user.profile
            )
        
            messages.success(request, 'Record uploaded successfully.')
            return redirect('nurse_dashboard')  # Replace with the name of the view to redirect after creation

        return render(request, 'core/create_observation_chart.html', {'admission': admission})
    
    return redirect('dashboard')

@login_required(login_url='login')
def update_observation_chart(request, chart_id):
    if request.user.profile.role in ['Nurse', 'Admin']:
        observation_chart = get_object_or_404(ObservationChart, id=chart_id)

        if request.method == 'POST':
            observation_chart.temperature = request.POST.get('temperature')
            observation_chart.pulse = request.POST.get('pulse')
            observation_chart.respiratory_rate = request.POST.get('respiratory_rate')
            observation_chart.blood_pressure = request.POST.get('blood_pressure')
            observation_chart.oxygen_saturation = request.POST.get('oxygen_saturation')
            observation_chart.remark = request.POST.get('remark')

            observation_chart.save()
            messages.success(request, "Observation Chart updated successfully.")
            return redirect('nurse_dashboard')

        return render(request, 'core/update_observation_chart.html', {
            'observation_chart': observation_chart
        })

    return redirect('dashboard')



@login_required(login_url='login')
def delete_observation_chart(request, chart_id):
    if request.user.profile.role in ['Nurse', 'Admin']:
        observation_chart = get_object_or_404(ObservationChart, id=chart_id)
        observation_chart.delete()
        messages.success(request, "Observation Chart deleted successfully.")
        return redirect('nurse_dashboard')

    return redirect('dashboard')



# ======================= Fluid Chart ==============================
@login_required(login_url='login')
def create_fluid_chart(request, admission_id):

    if request.user.profile.role == 'Nurse' or request.user.profile.role == 'Admin':
        admission = get_object_or_404(Admission, id=admission_id)

        if request.method == 'POST':
            intake_type = request.POST.get('intake_type')
            intake_amount = request.POST.get('intake_amount')
            intake_route = request.POST.get('intake_route')
            output_type = request.POST.get('output_type')
            output_amount = request.POST.get('output_amount')
            output_remark = request.POST.get('output_remark')


    #        Create the Fluid Chart instance
            fluidChart = FluidChart.objects.create(
                admission=admission,
                intake_type=intake_type,
                intake_amount=intake_amount,
                intake_route=intake_route,
                output_type=output_type,
                output_amount=output_amount,
                output_remark=output_remark,
                recorded_by=request.user.profile
            )
        
            messages.success(request, 'Record uploaded successfully.')
            return redirect('nurse_dashboard')  # Replace with the name of the view to redirect after creation
        return render(request, 'core/create_fluid_chart.html', {'admission': admission})
    
    return redirect('dashboard')




@login_required(login_url='login')
def update_fluid_chart(request, chart_id):
    if request.user.profile.role in ['Nurse', 'Admin']:
        fluid_chart = get_object_or_404(FluidChart, id=chart_id)

        if request.method == 'POST':
            fluid_chart.intake_type = request.POST.get('intake_type')
            fluid_chart.intake_amount = request.POST.get('intake_amount')
            fluid_chart.intake_route = request.POST.get('intake_route')
            fluid_chart.output_type = request.POST.get('output_type')
            fluid_chart.output_amount = request.POST.get('output_amount')
            fluid_chart.output_remark = request.POST.get('output_remark')

            fluid_chart.save()
            messages.success(request, "Fluid Chart updated successfully.")
            return redirect('nurse_dashboard')

        return render(request, 'core/update_fluid_chart.html', {'fluid_chart': fluid_chart})

    return redirect('dashboard')


@login_required(login_url='login')
def delete_fluid_chart(request, chart_id):
    if request.user.profile.role in ['Nurse', 'Admin']:
        fluid_chart = get_object_or_404(FluidChart, id=chart_id)
        fluid_chart.delete()
        messages.success(request, "Fluid Chart deleted successfully.")
        return redirect('nurse_dashboard')

    return redirect('dashboard')



# =============== Seizure Chart ======================

@login_required(login_url='login')
def create_seizure_chart(request, admission_id):

    if request.user.profile.role == 'Nurse' or request.user.profile.role == 'Admin':
        admission = get_object_or_404(Admission, id=admission_id)

        if request.method == 'POST':
            duration = request.POST.get('duration')
            description = request.POST.get('description')

            # Create the Seizure Chart instance
            seizure_chart = SeizureChart.objects.create(
                admission=admission,
                duration=duration,
                description=description,
                recorded_by=request.user.profile
            )
        
            messages.success(request, 'Seizure chart created successfully.')
            return redirect('nurse_dashboard')  # Replace with the name of the view to redirect after creation

        return render(request, 'core/create_seizure_chart.html', {'admission': admission})
    return redirect('dashboard')

@login_required(login_url='login')
def update_seizure_chart(request, chart_id):
    if request.user.profile.role in ['Nurse', 'Admin']:
        seizure_chart = get_object_or_404(SeizureChart, id=chart_id)

        if request.method == 'POST':
            seizure_chart.duration = request.POST.get('duration')
            seizure_chart.description = request.POST.get('description')

            seizure_chart.save()
            messages.success(request, "Seizure Chart updated successfully.")
            return redirect('nurse_dashboard')

        return render(request, 'core/update_seizure_chart.html', {'seizure_chart': seizure_chart})

    return redirect('dashboard')

@login_required(login_url='login')
def delete_seizure_chart(request, chart_id):
    if request.user.profile.role in ['Nurse', 'Admin']:
        seizure_chart = get_object_or_404(SeizureChart, id=chart_id)
        seizure_chart.delete()
        messages.success(request, "Seizure Chart deleted successfully.")
        return redirect('nurse_dashboard')

    return redirect('dashboard')


#=============== Nurse Handover =====================


@login_required(login_url='login')
def create_nurse_handover(request):
    if request.method == 'POST':
        notes = request.POST.get('notes')
        recorded_by = request.user.profile  # Assuming the user profile is related to `Profile`

        if notes:
            NurseHandover.objects.create(
                notes=notes,
                recorded_by=recorded_by
            )
            messages.success(request, "Nurse handover note created successfully.")
            return redirect('list_nurse_handover')  # Replace with your desired redirect URL
        else:
            messages.error(request, "Notes field cannot be empty.")
    
    return render(request, 'core/create_nurse_handover.html')

@login_required(login_url='login')
def update_nurse_handover(request, pk):
    try:
        handover_note = NurseHandover.objects.get(id=pk)
    except NurseHandover.DoesNotExist:
        messages.error(request, "Nurse handover note not found.")
        return redirect('list_nurse_handover')

    if request.method == 'POST':
        notes = request.POST.get('notes')

        if notes:
            handover_note.notes = notes
            handover_note.save()
            messages.success(request, "Nurse handover note updated successfully.")
            return redirect('list_nurse_handover')  # Replace with your desired redirect URL
        else:
            messages.error(request, "Notes field cannot be empty.")

    return render(request, 'core/update_nurse_handover.html', {
        'handover_note': handover_note
    })

@login_required(login_url='login')
def delete_nurse_handover(request, pk):
    try:
        handover_note = NurseHandover.objects.get(id=pk)
    except NurseHandover.DoesNotExist:
        messages.error(request, "Nurse handover note not found.")
        return redirect('list_nurse_handover')

    if request.method == 'POST':
        handover_note.delete()
        messages.success(request, "Nurse handover note deleted successfully.")
        return redirect('list_nurse_handover')  # Replace with your desired redirect URL

    return render(request, 'core/delete_nurse_handover.html', {
        'handover_note': handover_note
    })


@login_required(login_url='login')
def list_nurse_handover(request):
    if request.user.profile.role in ['Admin', 'Nurse']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        handover_notes = NurseHandover.objects.filter(recorded_at__date__range=[start_date, end_date]).order_by('-recorded_at')

        return render(request, 'core/list_nurse_handover.html', {
            'handover_notes': handover_notes,
            'start_date': start_date,
            'end_date': end_date,
        })
    return redirect('dashboard')




@login_required(login_url='login')
def view_admission(request, admission_id):

    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor']:
        # Ensure the user is authorized to view the admission details
        admission = get_object_or_404(Admission, id=admission_id)

        # Get related prescriptions, investigation requests, observation charts, fluid charts, seizure charts, Down's score charts, doctor's notes, and diagnoses
        prescriptions = Prescription.objects.filter(admission=admission)
        investigation_requests = InvestigationRequest.objects.filter(admission=admission)
        observation_charts = ObservationChart.objects.filter(admission=admission)
        fluid_charts = FluidChart.objects.filter(admission=admission)
        seizure_charts = SeizureChart.objects.filter(admission=admission)
        down_score_charts = DownScoreChart.objects.filter(admission=admission)
        doctor_notes = DoctorNote.objects.filter(admission=admission).order_by('-created_at')
        diagnoses = Diagnosis.objects.filter(admission=admission).order_by('-created_at')

        context = {
            'admission': admission,
            'prescriptions': prescriptions,
            'investigation_requests': investigation_requests,
            'observation_charts': observation_charts,
            'fluid_charts': fluid_charts,
            'seizure_charts': seizure_charts,
            'down_score_charts': down_score_charts,
            'doctor_notes': doctor_notes,
            'diagnoses': diagnoses,
        }

        return render(request, 'core/view_admission.html', context)

    return redirect('dashboard')



# =============== Down's Score Chart ========================

@login_required(login_url='login')
def create_down_score_chart(request, admission_id):

    if request.user.profile.role == 'Nurse' or request.user.profile.role == 'Admin':
        admission = get_object_or_404(Admission, id=admission_id)

        if request.method == 'POST':
            respiratory_rate = int(request.POST.get('respiratory_rate'))
            cyanosis = int(request.POST.get('cyanosis'))
            retraction = int(request.POST.get('retraction'))
            grunting = int(request.POST.get('grunting'))
            air_entry = int(request.POST.get('air_entry'))

            # Calculate the total score
            total = respiratory_rate + cyanosis + retraction + grunting + air_entry

            # Create the Down Score Chart instance
            down_score_chart = DownScoreChart.objects.create(
                admission=admission,
                respiratory_rate=respiratory_rate,
                cyanosis=cyanosis,
                retraction=retraction,
                grunting=grunting,
                air_entry=air_entry,
                total=total,
                recorded_by=request.user.profile
            )

            messages.success(request, "Down's Score Chart created successfully.")
            return redirect('nurse_dashboard')  # Replace with the name of the view to redirect after creation

        return render(request, 'core/create_down_score_chart.html', {'admission': admission})
    return redirect('dashboard')



@login_required(login_url='login')
def update_down_score_chart(request, chart_id):
    if request.user.profile.role in ['Nurse', 'Admin']:
        down_score_chart = get_object_or_404(DownScoreChart, id=chart_id)

        if request.method == 'POST':
            down_score_chart.respiratory_rate = int(request.POST.get('respiratory_rate'))
            down_score_chart.cyanosis = int(request.POST.get('cyanosis'))
            down_score_chart.retraction = int(request.POST.get('retraction'))
            down_score_chart.grunting = int(request.POST.get('grunting'))
            down_score_chart.air_entry = int(request.POST.get('air_entry'))

            # Calculate the total score
            down_score_chart.total = (
                down_score_chart.respiratory_rate +
                down_score_chart.cyanosis +
                down_score_chart.retraction +
                down_score_chart.grunting +
                down_score_chart.air_entry
            )

            down_score_chart.save()
            messages.success(request, "Down's Score Chart updated successfully.")
            return redirect('nurse_dashboard')

        return render(request, 'core/update_down_score_chart.html', {'down_score_chart': down_score_chart})
    
    return redirect('dashboard')


@login_required(login_url='login')
def delete_down_score_chart(request, chart_id):
    if request.user.profile.role in ['Nurse', 'Admin']:
        down_score_chart = get_object_or_404(DownScoreChart, id=chart_id)
        down_score_chart.delete()
        messages.success(request, "Down's Score Chart deleted successfully.")
        return redirect('nurse_dashboard')

    return redirect('dashboard')


# ================ Vital Sign=================

@login_required(login_url='login')
def add_vital_sign(request, appointment_id):

    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        appointment = get_object_or_404(Appointment, id=appointment_id)

        if request.method == 'POST':
            height = request.POST.get('height')
            weight = request.POST.get('weight')
            blood_pressure = request.POST.get('blood_pressure')
            temperature = request.POST.get('temperature')
            pulse = request.POST.get('pulse')
            respiratory_rate = request.POST.get('respiratory_rate')
            oxygen_saturation = request.POST.get('oxygen_saturation')

            # Create and save the vital sign record
            vital_sign = VitalSign.objects.create(
                appointment=appointment,
                height=height,
                weight=weight,
                blood_pressure=blood_pressure,
                temperature=temperature,
                pulse=pulse,
                respiratory_rate=respiratory_rate,
                oxygen_saturation=oxygen_saturation,
            )
            return redirect('view_appointment', appointment_id=appointment.id)

        return render(request, 'core/add_vital_sign.html', {'appointment': appointment})
    return redirect('dashboard')


@login_required(login_url='login')
def update_vital_sign(request, appointment_id, vital_sign_id):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        appointment = get_object_or_404(Appointment, id=appointment_id)
        vital_sign = get_object_or_404(VitalSign, id=vital_sign_id)

        if request.method == 'POST':
            vital_sign.height = request.POST.get('height')
            vital_sign.weight = request.POST.get('weight')
            vital_sign.blood_pressure = request.POST.get('blood_pressure')
            vital_sign.temperature = request.POST.get('temperature')
            vital_sign.pulse = request.POST.get('pulse')
            vital_sign.respiratory_rate = request.POST.get('respiratory_rate')
            vital_sign.oxygen_saturation = request.POST.get('oxygen_saturation')
            
            # Save the updated vital sign record
            vital_sign.save()
            messages.success(request, "Vital signs updated successfully.")
            return redirect('view_appointment', appointment_id=appointment.id)

        return render(request, 'core/update_vital_sign.html', {
            'appointment': appointment,
            'vital_sign': vital_sign
        })
    return redirect('dashboard')


@login_required(login_url='login')
def delete_vital_sign(request, appointment_id, vital_sign_id):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        vital_sign = get_object_or_404(VitalSign, id=vital_sign_id)
        vital_sign.delete()
        messages.success(request, "Vital signs deleted successfully.")
        return redirect('view_appointment', appointment_id=appointment_id)

    return redirect('dashboard')

# ============= Doctors ==============================
@login_required(login_url='login')
def doctor_dashboard(request):

    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        appointments = Appointment.objects.filter(appointment_date__date__range=[start_date, end_date]).order_by('-appointment_date', '-status')

        return render(request, 'core/doctor_dashboard.html', {
            'appointments': appointments,
            'start_date': start_date,
            'end_date': end_date,
        })
    return redirect('dashboard')

# ============= Doctors ==============================

@login_required(login_url='login')
def admissions(request):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        admissions = Admission.objects.filter(admission_date__date__range=[start_date, end_date]).order_by('-created_at')

        return render(request, 'core/admissions.html', {
            'admissions': admissions,
            'start_date': start_date,
            'end_date': end_date,
        })
    return redirect('dashboard')


@login_required(login_url='login')
def nurse_dashboard(request):
    if request.user.profile.role in ['Nurse', 'Admin']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        admissions = Admission.objects.filter(admission_date__date__range=[start_date, end_date]).order_by('-created_at')

        return render(request, 'core/nurse_dashboard.html', {
            'admissions': admissions,
            'start_date': start_date,
            'end_date': end_date,
        })
    return redirect('dashboard')


@login_required(login_url='login')
def add_prescriptions(request, appointment_id=None, admission_id=None):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        context = {}
        if appointment_id:
            appointment = get_object_or_404(Appointment, id=appointment_id)
            context['appointment'] = appointment
            has_diagnosis = Diagnosis.objects.filter(appointment=appointment).exists()
        elif admission_id:
            admission = get_object_or_404(Admission, id=admission_id)
            context['admission'] = admission
            has_diagnosis = Diagnosis.objects.filter(admission=admission).exists()
        else:
            return redirect('dashboard')

        if not has_diagnosis:
            messages.error(request, "Please record a diagnosis first, then click Prescription again to add medication.")
            if appointment_id:
                return redirect('create_diagnosis_appointment', appointment_id=appointment_id)
            else:
                return redirect('create_diagnosis_admission', admission_id=admission_id)

        medications = Medication.objects.all()  # Fetch all available medications

        if request.method == 'POST':
            medication_ids = request.POST.getlist('medication_id')
            dosages = request.POST.getlist('dosage')
            instructions = request.POST.getlist('instructions')

            for i in range(len(medication_ids)):
                medication_id = medication_ids[i]
                dosage = dosages[i]
                instruction = instructions[i]

                if medication_id and dosage:
                    medication = Medication.objects.get(id=medication_id)
                    Prescription.objects.create(
                        appointment=appointment if appointment_id else None,
                        admission=admission if admission_id else None,
                        medication=medication,
                        dosage=dosage,
                        instructions=instruction
                    )

            if appointment_id:
                return redirect('doctor_dashboard')
            else:
                return redirect('admissions')

        context['medications'] = medications
        return render(request, 'core/add_prescriptions.html', context)
    
    return redirect('dashboard')


@login_required(login_url='login')
def create_investigation_requests(request, appointment_id=None, admission_id=None):
    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        context = {}
        if appointment_id:
            appointment = get_object_or_404(Appointment, id=appointment_id)
            context['appointment'] = appointment
        elif admission_id:
            admission = get_object_or_404(Admission, id=admission_id)
            context['admission'] = admission
        else:
            return redirect('dashboard')

        investigations = Investigation.objects.all()  # Fetch all available investigations

        if request.method == 'POST':
            selected_investigation_ids = request.POST.getlist('investigation_id')  # Get the list of selected investigations

            if selected_investigation_ids:
                invoice = Invoice.objects.create(patient=appointment.patient if appointment_id else admission.patient)
                total_amount = 0

                for investigation_id in selected_investigation_ids:
                    if investigation_id:  # Ensure the investigation_id is not empty or None
                        try:
                            investigation = Investigation.objects.get(id=investigation_id)
                            investigation_request = InvestigationRequest.objects.create(
                                appointment=appointment if appointment_id else None,
                                admission=admission if admission_id else None,
                                investigation=investigation,
                                patient=appointment.patient if appointment_id else admission.patient
                            )

                            # Create invoice item
                            invoice_item = InvoiceItem.objects.create(
                                invoice=invoice,
                                
                                investigation_request=investigation_request,
                                quantity=1,
                                unit_price=investigation.price,
                                total_price=investigation.price
                            )
                            total_amount += invoice_item.total_price
                        except Investigation.DoesNotExist:
                            # Handle case where the investigation does not exist
                            messages.error(request, "Selected investigation does not exist.")
                            return redirect('create_investigation_requests', appointment_id=appointment_id, admission_id=admission_id)

                invoice.total_amount = total_amount
                invoice.save()

                messages.success(request, f"Invoice #{invoice.invoice_number} created successfully.")
                if appointment_id:
                    return redirect('doctor_dashboard')
                else:
                    return redirect('admissions')

        context['investigations'] = investigations
        return render(request, 'core/create_investigation_requests.html', context)

    return redirect('dashboard')





@login_required(login_url='login')
def create_admission(request, appointment_id):

    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        appointment = get_object_or_404(Appointment, id=appointment_id)

        if request.method == 'POST':
            reason = request.POST.get('reason')
            ward = request.POST.get('ward')

            Admission.objects.create(
                patient=appointment.patient,
                doctor=appointment.doctor,
                reason=reason,
                ward=ward,
            )
            return redirect('admissions')

        return render(request, 'core/create_admission.html', {'appointment': appointment})
    return redirect('dashboard')

@login_required(login_url='login')
def complete_appointment(request, appointment_id):

    if request.user.profile.role in ['Nurse', 'Admin', 'Doctor', 'Receptionist']:
        # Get the appointment and mark it as completed
        appointment = get_object_or_404(Appointment, id=appointment_id)
        appointment.status = 'Completed'
        appointment.save()

    # Redirect back to the doctor's dashboard or another appropriate page
    return redirect('doctor_dashboard')


@login_required(login_url='login')
def delete_appointment(request, appointment_id):
    Appointment.objects.get(id=appointment_id).delete()
    return redirect('appointment_list')

# ======================= Investigation Result =====================

@login_required(login_url='login')
def view_investigation_result(request, request_id):
    investigation_request = get_object_or_404(InvestigationRequest, id=request_id)
    investigation_result = get_object_or_404(InvestigationResult, request=investigation_request)
    print(investigation_result)

    context = {
        'investigation_request': investigation_request,
        'investigation_result': investigation_result,
    }

    return render(request, 'core/view_investigation_result.html', context)

# =======================Doctors Note

@login_required(login_url='login')
def view_notes(request, appointment_id=None, admission_id=None):
    notes = None
    if appointment_id:
        appointment = get_object_or_404(Appointment, id=appointment_id)
        notes = DoctorNote.objects.filter(appointment=appointment).order_by('-created_at')
        context_type = 'Appointment'
    elif admission_id:
        admission = get_object_or_404(Admission, id=admission_id)
        notes = DoctorNote.objects.filter(admission=admission).order_by('-created_at')
        context_type = 'Admission'
    else:
        return redirect('dashboard')

    if request.method == 'POST':
        note_content = request.POST.get('note')
        if note_content:
            DoctorNote.objects.create(
                appointment=appointment if appointment_id else None,
                admission=admission if admission_id else None,
                doctor=request.user,
                note=note_content
            )
            if appointment_id:
                return redirect('view_appointment_notes', appointment_id=appointment.id)
            elif admission_id:
                return redirect('view_admission_notes', admission_id=admission.id)

    context = {
        'notes': notes,
        'context_type': context_type,
        'appointment': appointment if appointment_id else None,
        'admission': admission if admission_id else None,
    }

    return render(request, 'core/view_notes.html', context)


# =======================Pharmacy====================






@login_required(login_url='login')
def pharmacy_dashboard(request):
    if request.user.profile.role in ['Pharmacist', 'Admin', 'PharmacyHead']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        appointments_with_prescriptions = Appointment.objects.filter(
            prescription__isnull=False, appointment_date__date__range=[start_date, end_date]
        ).order_by('-appointment_date').distinct()

        # Calculate invoicing status
        appointments_with_invoicing_status = []
        for appointment in appointments_with_prescriptions:
            prescriptions = appointment.prescription_set.all().order_by('-created_at')
            all_invoiced = all(prescription.is_invoiced for prescription in prescriptions)
            appointments_with_invoicing_status.append({
                'appointment': appointment,
                'all_invoiced': all_invoiced,
            })

        return render(request, 'core/pharmacy_dashboard.html', {
            'appointments_with_invoicing_status': appointments_with_invoicing_status,
            'start_date': start_date,
            'end_date': end_date,
        })

    return redirect('dashboard')




@login_required(login_url='login')
def create_invoice(request, appointment_id):
    if request.user.profile.role in ['Pharmacist', 'Admin', 'Cashier']:
        appointment = get_object_or_404(Appointment, id=appointment_id)
        prescriptions = Prescription.objects.filter(appointment=appointment, is_invoiced=False)

        if request.method == 'POST':
            selected_medications = request.POST.getlist('medication_id')
            quantities = request.POST.getlist('quantity')
            deletions = request.POST.getlist('delete')

            invoice = Invoice.objects.create(patient=appointment.patient)
            total_amount = 0

            for index, med_id in enumerate(selected_medications):
                if str(index) not in deletions:  # Only include if not marked for deletion
                    medication = Medication.objects.get(id=med_id)
                    unit_price = medication.unit_price or 0  # Use 0 if unit_price is None
                    quantity = int(quantities[index])
                    total_price = unit_price * quantity

                    # Create InvoiceItem with total_price
                    invoice_item = InvoiceItem.objects.create(
                        invoice=invoice,
                        medication=medication,
                        quantity=quantity,
                        unit_price=unit_price,
                        total_price=total_price  # Ensure total price is calculated correctly
                    )
                    total_amount += total_price

                    # Update the prescription status
                    prescription = Prescription.objects.get(appointment=appointment, medication=medication)
                    prescription.is_invoiced = True
                    prescription.save()

            invoice.total_amount = total_amount
            invoice.save()
            messages.success(request, f"Invoice #{invoice.invoice_number} created successfully.")
            return redirect('pharmacy_dashboard')

        all_medications = Medication.objects.all()

        return render(request, 'core/create_invoice.html', {
            'appointment': appointment,
            'prescriptions': prescriptions,
            'all_medications': all_medications,
        })
    return redirect('dashboard')


# @login_required(login_url='login')
# def create_invoice(request, appointment_id):
#     appointment = get_object_or_404(Appointment, id=appointment_id)
#     prescriptions = Prescription.objects.filter(appointment=appointment, is_invoiced=False)
#     available_stock = {stock.medication.id: stock for stock in PharmacyStock.objects.filter(quantity__gt=0)}

#     if request.method == 'POST':
#         selected_medications = request.POST.getlist('medication_id')
#         quantities = request.POST.getlist('quantity')
#         deletions = request.POST.getlist('delete')

#         invoice = Invoice.objects.create(patient=appointment.patient)
#         total_amount = 0

#         for index, med_id in enumerate(selected_medications):
#             if str(index) not in deletions:  # Only include if not marked for deletion
#                 medication = Medication.objects.get(id=med_id)
#                 stock = available_stock.get(int(med_id))

#                 if stock and stock.quantity >= int(quantities[index]):
#                     invoice_item = InvoiceItem.objects.create(
#                         invoice=invoice,
#                         medication=medication,
#                         quantity=int(quantities[index])
#                     )
#                     total_amount += invoice_item.total_price
#                     stock.quantity -= int(quantities[index])
#                     stock.save()
                    
#                     # Update the prescription status
#                     prescription = Prescription.objects.get(appointment=appointment, medication=medication)
#                     prescription.is_invoiced = True
#                     prescription.save()
#                 else:
#                     messages.error(request, f"Not enough stock for {medication.name}")

#         invoice.total_amount = total_amount
#         invoice.save()
#         messages.success(request, f"Invoice #{invoice.invoice_number} created successfully.")
#         return redirect('pharmacy_dashboard')

#     all_medications = Medication.objects.all()

#     return render(request, 'core/create_invoice.html', {
#         'appointment': appointment,
#         'prescriptions': prescriptions,
#         'available_stock': available_stock,
#         'all_medications': all_medications,
#     })






@login_required(login_url='login')
def sell_medications(request):
    if request.user.profile.role in ['Pharmacist', 'Admin', 'Cashier']:
        medications = Medication.objects.all()
        patients = Patient.objects.all()  # Fetch all patients

        if request.method == 'POST':
            patient_id = request.POST.get('patient_id')
            selected_medications = request.POST.getlist('medication_id')
            quantities = request.POST.getlist('quantity')

            if not selected_medications or not quantities:
                messages.error(request, "No medications selected.")
                return redirect('sell_medications')

            patient = get_object_or_404(Patient, id=patient_id) if patient_id else None
            invoice = Invoice.objects.create(patient=patient, created_by=request.user.profile)  # Create invoice with the selected patient
            total_amount = 0

            for med_id, qty in zip(selected_medications, quantities):
                medication = Medication.objects.get(id=med_id)
                qty = int(qty)

                if medication.unit_price is None:
                    messages.error(request, f"Medication {medication.name} has no unit price set.")
                    return redirect('sell_medications')

                invoice_item = InvoiceItem.objects.create(
                    invoice=invoice,
                    medication=medication,
                    quantity=qty,
                    unit_price=medication.unit_price,
                    total_price=medication.unit_price * qty
                )
                total_amount += invoice_item.total_price

            invoice.total_amount = total_amount
            invoice.save()
            messages.success(request, f"Invoice #{invoice.invoice_number} created successfully.")
            return redirect('manage_medication_invoices')

        return render(request, 'core/sell_medications.html', {
            'medications': medications,
            'patients': patients,
        })

    return redirect('dashboard')




# @login_required(login_url='login')
# def sell_medications(request):

#     if request.user.profile.role == 'Pharmacist' or request.user.profile.role == 'Admin' or request.user.profile.role == 'Cashier':
#         medications = Medication.objects.all()
#         patients = Patient.objects.all()  # Fetch all patients
#         available_stock = {stock.medication.id: stock for stock in PharmacyStock.objects.filter(quantity__gt=0)}

#         if request.method == 'POST':
#             patient_id = request.POST.get('patient_id')
#             selected_medications = request.POST.getlist('medication_id')
#             quantities = request.POST.getlist('quantity')

#             if not selected_medications or not quantities:
#                 messages.error(request, "No medications selected.")
#                 return redirect('sell_medications')

#             patient = get_object_or_404(Patient, id=patient_id) if patient_id else None
#             invoice = Invoice.objects.create(patient=patient)  # Create invoice with the selected patient
#             total_amount = 0

#             for med_id, qty in zip(selected_medications, quantities):
#                 medication = Medication.objects.get(id=med_id)
#                 stock = available_stock.get(int(med_id))

#                 if stock and stock.quantity >= int(qty):
#                     invoice_item = InvoiceItem.objects.create(
#                         invoice=invoice,
#                         medication=medication,
#                         quantity=int(qty)
#                     )
#                     total_amount += invoice_item.total_price
#                     stock.quantity -= int(qty)
#                     stock.save()
#                 else:
#                     messages.error(request, f"Not enough stock for {medication.name}")

#             invoice.total_amount = total_amount
#             invoice.save()
#             messages.success(request, f"Invoice #{invoice.invoice_number} created successfully.")
#             return redirect('manage_medication_invoices')

#         return render(request, 'core/sell_medications.html', {
#             'medications': medications,
#             'patients': patients,
#             'available_stock': available_stock,
#         })

#     return redirect('dashboard')



# ================ Cashier Dashboard==================

@login_required(login_url='login')
def cashier_dashboard(request):
    if request.user.profile.role in ['Cashier', 'Admin', 'Accountant']:
        # Daily totals by cashier, now including Wallet payments
        daily_cashier_totals = Payment.objects.filter(
            payment_date__date=datetime.today().date()
        ).values(
            cashier_name=F('created_by__user__username'),
            payment_method_name=F('payment_method')
        ).annotate(
            total_amount=Sum('amount')
        ).order_by('cashier_name')

        cashier_data = {}
        for entry in daily_cashier_totals:
            cashier = entry['cashier_name']
            payment_method = entry['payment_method_name']
            total_amount = entry['total_amount']
            
            if cashier not in cashier_data:
                cashier_data[cashier] = {
                    'pos': 0,
                    'cash': 0,
                    'bank_transfer': 0,
                    'insurance': 0,
                    'wallet': 0,  # For direct wallet payments
                    'other': 0,
                    'total': 0,
                    'refunds': 0,
                    'wallet_credits': 0,  # For wallet credits summary
                    'net_total': 0,  # Net total after refunds
                }

            if payment_method.lower() == 'pos':
                cashier_data[cashier]['pos'] += total_amount
            elif payment_method.lower() == 'cash':
                cashier_data[cashier]['cash'] += total_amount
            elif payment_method.lower() == 'bank transfer':
                cashier_data[cashier]['bank_transfer'] += total_amount
            elif payment_method.lower() == 'insurance':
                cashier_data[cashier]['insurance'] += total_amount
            elif payment_method.lower() == 'wallet':
                cashier_data[cashier]['wallet'] += total_amount  # Add wallet payments here
            else:
                cashier_data[cashier]['other'] += total_amount

            cashier_data[cashier]['total'] += total_amount

        # Handle refunds as usual
        daily_refunds = Refund.objects.filter(
            created_at__date=datetime.today().date()
        ).values(
            cashier_name=F('payment__created_by__user__username')
        ).annotate(
            total_refunds=Sum('amount')
        ).order_by('cashier_name')

        for entry in daily_refunds:
            cashier = entry['cashier_name']
            total_refunds = entry['total_refunds']
            if cashier in cashier_data:
                cashier_data[cashier]['refunds'] += total_refunds
                # Deduct refunds from the total to get the net total
                cashier_data[cashier]['net_total'] = cashier_data[cashier]['total'] - cashier_data[cashier]['refunds']

        # Wallet credits total per cashier and payment method
        daily_wallet_credits = Transaction.objects.filter(
            transaction_type='Credit',
            date__date=datetime.today().date()
        ).values(
            cashier_name=F('created_by__user__username'),
            transaction_method_name=F('transaction_method')
        ).annotate(
            total_credits=Sum('amount')
        ).order_by('cashier_name')

        for entry in daily_wallet_credits:
            cashier = entry['cashier_name']
            transaction_method = entry['transaction_method_name']
            total_credits = entry['total_credits']
            
            if cashier not in cashier_data:
                cashier_data[cashier] = {
                    'pos': 0,
                    'cash': 0,
                    'bank_transfer': 0,
                    'insurance': 0,
                    'wallet': 0,
                    'other': 0,
                    'total': 0,
                    'refunds': 0,
                    'wallet_credits': 0,
                    'net_total': 0,
                }

            # Add wallet credits to the respective payment method and to the wallet_credits summary
            if transaction_method and transaction_method.lower() == 'pos':
                cashier_data[cashier]['pos'] += total_credits
            elif transaction_method and transaction_method.lower() == 'cash':
                cashier_data[cashier]['cash'] += total_credits
            elif transaction_method and transaction_method.lower() == 'bank transfer':
                cashier_data[cashier]['bank_transfer'] += total_credits
            else:
                cashier_data[cashier]['other'] += total_credits  # Handle unknown methods if any

            # Add total credits to the wallet_credits field
            cashier_data[cashier]['wallet_credits'] += total_credits

            # Update total and net total
            cashier_data[cashier]['total'] += total_credits
            cashier_data[cashier]['net_total'] = cashier_data[cashier]['total'] - cashier_data[cashier]['refunds']

        # Cashiers only see their own totals; Admin/Accountant see everyone's
        if request.user.profile.role == 'Cashier':
            cashier_data = {
                cashier: data for cashier, data in cashier_data.items()
                if cashier == request.user.username
            }

        # Filtering logic for pending payments
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        # Subqueries replace the old per-invoice Payment/Refund aggregate loop (N+1 fix).
        paid_sq = (
            Payment.objects.filter(invoice=OuterRef('pk'))
            .values('invoice')
            .annotate(s=Sum('amount'))
            .values('s')
        )
        refunded_sq = (
            Refund.objects.filter(payment__invoice=OuterRef('pk'))
            .values('payment__invoice')
            .annotate(s=Sum('amount'))
            .values('s')
        )

        invoices_qs = (
            Invoice.objects.filter(created_at__date__range=[start_date, end_date])
            .select_related('patient', 'family')
            .prefetch_related('items__service', 'items__medication', 'items__investigation_request__investigation')
            .annotate(db_total_paid=Subquery(paid_sq), db_total_refunded=Subquery(refunded_sq))
            .order_by('-created_at')
        )

        invoice_data = []
        for invoice in invoices_qs:
            total_paid = invoice.db_total_paid or 0
            total_refunded = invoice.db_total_refunded or 0

            if invoice.patient:
                patient_name = f"{invoice.patient.first_name} {invoice.patient.last_name} [ {invoice.patient.patient_file_number} ]"
            elif invoice.family:
                patient_name = f"{invoice.family.name} [ Family ]"
            elif invoice.custom_name:
                patient_name = invoice.custom_name
            else:
                patient_name = "N/A"

            invoice_data.append({
                'invoice': invoice,
                'total_paid': total_paid,
                'refund': total_refunded,
                'balance_due': invoice.total_amount - total_paid + total_refunded,
                'items': invoice.items.all(),
                'patient_name': patient_name,
            })

        context = {
            'invoice_data': invoice_data,
            'cashier_data': cashier_data,
            'start_date': start_date,
            'end_date': end_date,
        }

        return render(request, 'core/cashier_dashboard.html', context)

    return redirect('dashboard')


@login_required(login_url='login')
def merge_invoices(request):
    if request.user.profile.role not in ['Admin', 'Accountant']:
        return redirect('dashboard')

    if request.method == 'POST':
        invoice_ids = request.POST.getlist('invoices')

        if len(invoice_ids) < 2:
            messages.error(request, "Please select at least two invoices to merge.")
            return redirect('cashier_dashboard')

        invoices = Invoice.objects.filter(id__in=invoice_ids)

        # Get distinct patient IDs and family IDs
        patient_ids = invoices.values_list('patient_id', flat=True).distinct()
        family_ids = invoices.values_list('family_id', flat=True).distinct()

        # Check if all selected invoices are for the same patient or family
        if len(patient_ids) > 1 or len(family_ids) > 1:
            messages.error(request, "Invoices must be for the same patient or family.")
            return redirect('cashier_dashboard')

        # Create a new merged invoice
        with transaction.atomic():
            merged_invoice = Invoice.objects.create(
                patient=invoices.first().patient,
                family=invoices.first().family,
                total_amount=0  # Will be updated later
            )

            # Move items from old invoices to the new invoice
            total_amount = 0
            for invoice in invoices:
                for item in invoice.items.all():
                    item.invoice = merged_invoice
                    item.save()
                    total_amount += item.total_price

            # Update the total amount for the merged invoice
            merged_invoice.total_amount = total_amount
            merged_invoice.save()

            # Delete the old invoices
            invoices.delete()

        messages.success(request, "Invoices merged successfully.")
        return redirect('cashier_dashboard')

    return redirect('dashboard')


@login_required(login_url='login')
def delete_invoices(request):
    if request.user.profile.role not in ['Admin', 'Accountant']:
        return redirect('dashboard')

    if request.method == 'POST':
        invoice_ids = request.POST.getlist('invoices')

        if not invoice_ids:
            messages.error(request, "Please select at least one invoice to delete.")
            return redirect('cashier_dashboard')

        invoices = Invoice.objects.filter(id__in=invoice_ids)

        # Get distinct patient IDs and family IDs
        patient_ids = invoices.values_list('patient_id', flat=True).distinct()
        family_ids = invoices.values_list('family_id', flat=True).distinct()

        # Ensure invoices are for the same patient or family
        if len(patient_ids) > 1 or len(family_ids) > 1:
            messages.error(request, "Invoices must be for the same patient or family.")
            return redirect('cashier_dashboard')

        # Ensure all invoices are unpaid
        unpaid_invoices = invoices.filter(is_paid=False)
        if unpaid_invoices.count() != len(invoice_ids):
            messages.error(request, "Only unpaid invoices can be deleted.")
            return redirect('cashier_dashboard')

        # Delete the invoices
        with transaction.atomic():
            invoices.delete()

        messages.success(request, "Selected invoices deleted successfully.")
        return redirect('cashier_dashboard')

    return redirect('dashboard')


@login_required(login_url='login')
def update_invoice_payment(request, invoice_id):
    if request.user.profile.role not in ['Cashier', 'Admin', 'Accountant']:
        return redirect('dashboard')

    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # Total paid so far
    total_paid = Payment.objects.filter(invoice=invoice).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Total refunded so far
    total_refunded = Refund.objects.filter(payment__invoice=invoice).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Net paid = Total paid - Total refunded
    net_paid = total_paid - total_refunded
    balance_due = invoice.total_amount - net_paid

    patient = invoice.patient
    family = patient.family if patient else None

    if request.method == 'POST':
        amount_str = request.POST.get('amount')
        payment_method = request.POST.get('payment_method')
        is_refund = request.POST.get('is_refund', 'false') == 'true'

        amount = Decimal(amount_str) if amount_str else balance_due

        if is_refund:
            # Handling refunds
            if amount <= 0:
                messages.error(request, "Refund amount must be positive.")
            elif amount > net_paid:
                messages.error(request, "Refund amount exceeds the net paid amount.")
            else:
                # Process refund
                payment = Payment.objects.filter(invoice=invoice).first()
                if payment:
                    Refund.objects.create(payment=payment, amount=amount, created_by=request.user.profile)
                    net_paid -= amount
                    balance_due = invoice.total_amount - net_paid

                    # Mark the invoice as unpaid if balance_due becomes positive
                    if balance_due > 0:
                        invoice.is_paid = False
                        invoice.save()

                    # Handle wallet refund logic
                    wallet = None
                    
                    if family and hasattr(family, 'wallet'):
                        wallet = family.wallet
                    elif patient and hasattr(patient, 'wallet'):
                        wallet = patient.wallet

                    if wallet:
                        # Credit the wallet balance
                        wallet.balance += amount
                        wallet.save()

                        # Create a transaction record for the refund
                        Transaction.objects.create(
                            wallet=wallet,
                            amount=amount,
                            transaction_type='Refund',
                            description=f'Refund for Invoice #{invoice.invoice_number}',
                            created_by=request.user.profile
                        )
                        messages.success(request, f"Refund of {amount} processed and credited to the wallet for Invoice #{invoice.invoice_number}.")
                    

                    messages.success(request, f"Refund of {amount} processed for Invoice #{invoice.invoice_number}.")
                    return redirect('view_invoice', invoice_id)
                else:
                    messages.error(request, "No payment record found for this invoice.")
        else:
            # Handling payments
            if amount <= 0:
                messages.error(request, "Payment amount must be positive.")
            elif amount > balance_due and payment_method != 'Wallet':
                messages.error(request, "Payment amount exceeds the balance due.")
            else:
                if payment_method == 'Wallet':
                    # Check if the patient has a family with a wallet, or a patient wallet
                    wallet = None
                    if family and hasattr(family, 'wallet'):
                        wallet = family.wallet
                    elif patient and hasattr(patient, 'wallet'):
                        wallet = patient.wallet

                    if wallet:
                        # Deduct the wallet balance
                        wallet.balance -= amount
                        wallet.save()

                        # Create a transaction record
                        Transaction.objects.create(
                            wallet=wallet,
                            amount=amount,
                            transaction_type='Debit',
                            description=f'Payment for Invoice #{invoice.invoice_number}',
                            created_by=request.user.profile
                        )

                        # Process payment as usual
                        Payment.objects.create(
                            invoice=invoice,
                            amount=amount,
                            payment_method='Wallet',
                            created_by=request.user.profile
                        )

                        net_paid += amount
                        balance_due = invoice.total_amount - net_paid

                        if balance_due <= 0:
                            invoice.is_paid = True
                            invoice.save()

                        messages.success(request, f"Invoice #{invoice.invoice_number} updated successfully using wallet.")
                        return redirect('view_invoice', invoice_id)
                    else:
                        messages.error(request, "No wallet found for the patient or family. Payment cannot be processed using wallet.")
                else:
                    # Process non-wallet payment
                    Payment.objects.create(
                        invoice=invoice,
                        amount=amount,
                        payment_method=payment_method,
                        created_by=request.user.profile
                    )

                    net_paid += amount
                    balance_due = invoice.total_amount - net_paid

                    if balance_due <= 0:
                        invoice.is_paid = True
                        invoice.save()

                    messages.success(request, f"Invoice #{invoice.invoice_number} updated successfully.")
                    return redirect('view_invoice', invoice_id)

    return render(request, 'core/update_invoice_payment.html', {
        'invoice': invoice,
        'total_paid': total_paid,
        'total_refunded': total_refunded,
        'net_paid': net_paid,
        'balance_due': balance_due
    })


@login_required(login_url='login')
def manage_medication_invoices(request):
    if request.user.profile.role in ['Admin', 'Cashier', 'Pharmacist', 'Accountant']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        invoices = Invoice.objects.filter(
            items__medication__isnull=False,
            created_at__date__range=[start_date, end_date]
        ).distinct().order_by('-created_at')

        return render(request, 'core/manage_medication_invoices.html', {
            'invoices': invoices,
            'start_date': start_date,
            'end_date': end_date,
        })

    return redirect('dashboard')







@login_required(login_url='login')
def update_invoice(request, invoice_id):

    if request.user.profile.role in ['Admin', 'Pharmacist', 'Accountant']:
        invoice = get_object_or_404(Invoice, id=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)

        if request.method == 'POST':
            selected_medications = request.POST.getlist('medication_id')
            quantities = request.POST.getlist('quantity')
            deletions = request.POST.getlist('delete')

            total_amount = 0

            for index, item in enumerate(invoice_items):
                if str(index) in deletions:
                    item.delete()  # Delete the invoice item if marked for deletion
                else:
                    medication = Medication.objects.get(id=selected_medications[index])

                    item.medication = medication
                    item.quantity = int(quantities[index])
                    item.save()
                    total_amount += item.total_price

            invoice.total_amount = total_amount
            invoice.save()

            if invoice.total_amount == 0:
                invoice.delete()

            messages.success(request, f"Invoice #{invoice.invoice_number} updated successfully.")
            return redirect('manage_medication_invoices')

        all_medications = Medication.objects.all()

        return render(request, 'core/update_invoice.html', {
            'invoice': invoice,
            'invoice_items': invoice_items,
            'all_medications': all_medications,
        })

    return redirect('dashboard')



# @login_required(login_url='login')
# def update_invoice(request, invoice_id):

#     if request.user.profile.role == 'Admin' or request.user.profile.role == 'Cashier':    
#         invoice = get_object_or_404(Invoice, id=invoice_id)
        
#         invoice_items = InvoiceItem.objects.filter(invoice=invoice)
#         available_stock = {stock.medication.id: stock for stock in PharmacyStock.objects.filter(quantity__gt=0)}

#         if request.method == 'POST':
#             selected_medications = request.POST.getlist('medication_id')
#             quantities = request.POST.getlist('quantity')
#             deletions = request.POST.getlist('delete')

#             total_amount = 0

#             for index, item in enumerate(invoice_items):
#                 if str(index) in deletions:
#                     # Restock the medication
#                     stock = available_stock.get(item.medication.id)
#                     stock.quantity += item.quantity
#                     stock.save()
#                     item.delete()
#                 else:
#                     medication = Medication.objects.get(id=selected_medications[index])
#                     stock = available_stock.get(medication.id)
#                     if stock and stock.quantity >= int(quantities[index]):
#                         stock.quantity += item.quantity  # Restock the old quantity
#                         stock.quantity -= int(quantities[index])  # Reduce the new quantity
#                         stock.save()

#                         item.medication = medication
#                         item.quantity = int(quantities[index])
#                         item.save()
#                         total_amount += item.total_price
#                     else:
#                         messages.error(request, f"Not enough stock for {medication.name}")

#             invoice.total_amount = total_amount
#             invoice.save()

#             messages.success(request, f"Invoice #{invoice.invoice_number} updated successfully.")
#             return redirect('manage_medication_invoices')

#         all_medications = Medication.objects.all()

#         return render(request, 'core/update_invoice.html', {
#             'invoice': invoice,
#             'invoice_items': invoice_items,
#             'available_stock': available_stock,
#             'all_medications': all_medications,
#         })

#     return redirect('dashboard')



# View to manage Medications
@login_required(login_url='login')
def manage_medications(request):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Pharmacist':

        medications = Medication.objects.all()
        if request.method == 'POST' and 'upload_file' in request.FILES:
            excel_file = request.FILES['upload_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)
            df = pd.read_excel(file_path)
            for index, row in df.iterrows():
                Medication.objects.create(
                    name=f"{row['name']} [{row['category']}]",
                    description=row['description'],
                    unit_price=row['price'],
                    supplier=Supplier.objects.get(id=row['supplier_id'])
                )

            messages.success(request, 'Medications imported successfully.')
            return redirect('manage_medications')
        return render(request, 'core/manage_medications.html', {'medications': medications})
    
    return redirect('dashboard')

# View to create or update a Medication
@login_required(login_url='login')
def medication_form(request, medication_id=None):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Pharmacist':
        if medication_id:
            medication = get_object_or_404(Medication, id=medication_id)
        else:
            medication = None

        if request.method == 'POST':
            name = request.POST['name']
            description = request.POST['description']
            unit_price = request.POST['unit_price']
            supplier_id = request.POST['supplier']
            supplier = get_object_or_404(Supplier, id=supplier_id)
            if medication:
                medication.name = name
                medication.description = description
                medication.unit_price = unit_price
                medication.supplier = supplier
                medication.save()
                messages.success(request, 'Medication updated successfully.')
            else:
                Medication.objects.create(name=name, description=description, unit_price=unit_price, supplier=supplier)
                messages.success(request, 'Medication created successfully.')
            return redirect('manage_medications')
        
        suppliers = Supplier.objects.all()
        return render(request, 'core/medication_form.html', {'medication': medication, 'suppliers': suppliers})

    return redirect('dashboard')

@login_required(login_url='login')
def delete_medication(request, medication_id):
    Medication.objects.get(id=medication_id).delete()
    messages.success(request, 'Medication deleted successfully.')
    return redirect('manage_medications')

# View to manage Pharmacy Stock
@login_required(login_url='login')
def manage_pharmacy_stock(request):

    if request.user.profile.role in ['Admin', 'Pharmacist', 'PharmacyHead']:
        pharmacy_stock = PharmacyStock.objects.all()
        if request.method == 'POST' and 'upload_file' in request.FILES:
            if request.user.profile.role not in ['Admin', 'PharmacyHead']:
                messages.error(request, 'Only the Pharmacy Head can add stock.')
                return redirect('manage_pharmacy_stock')

            excel_file = request.FILES['upload_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)
            df = pd.read_excel(file_path)
            for index, row in df.iterrows():
                stock = PharmacyStock.objects.create(
                    medication=Medication.objects.get(id=row['medication_id']),
                    quantity=row['quantity'],
                    unit=row['unit'],
                    expiry_date=row['expiry_date'],
                    batch_number=row.get('batch_number', ''),
                    manufacture_date=row.get('manufacture_date', None),
                    created_by=request.user.profile,
                )
                PharmacyStockTransaction.objects.create(
                    pharmacy_stock=stock,
                    transaction_type='Addition',
                    quantity=row['quantity'],
                    created_by=request.user.profile,
                )
            messages.success(request, 'Pharmacy stock imported successfully.')
            return redirect('manage_pharmacy_stock')
        total_amount = sum(i.amount() for i in pharmacy_stock)
        total_amount = f"₦{total_amount:,.2f}"
        return render(request, 'core/manage_pharmacy_stock.html', {'pharmacy_stock': pharmacy_stock, 'total_amount': total_amount})
    
    return redirect('dashboard')


@login_required(login_url='login')
def pharmacy_stock_form(request):
    if request.user.profile.role in ['Admin', 'PharmacyHead']:
        if request.method == 'POST':
            medications = request.POST.getlist('medications[]')  # Get the list of medication IDs
            quantities = request.POST.getlist('quantities[]')  # Get the list of quantities
            units = request.POST.getlist('units[]')  # Get the list of units
            unit_prices = request.POST.getlist('unit_prices[]')  # Get the list of units

            for i in range(len(medications)):
                medication_id = medications[i]
                quantity = quantities[i]
                unit = units[i]
                unit_price = unit_prices[i]

                medication = get_object_or_404(Medication, id=medication_id)

                # Create a new PharmacyStock entry for each selected medication
                stock = PharmacyStock.objects.create(
                    medication=medication,
                    quantity=quantity,
                    unit=unit,
                    unit_price=unit_price,
                    created_by=request.user.profile,

                )

                PharmacyStockTransaction.objects.create(
                    pharmacy_stock=stock,
                    transaction_type='Addition',
                    quantity=quantity,
                    created_by=request.user.profile,
                )

            messages.success(request, 'Pharmacy stock created successfully.')
            return redirect('manage_pharmacy_stock')
        
        medications = Medication.objects.all()
        return render(request, 'core/pharmacy_stock_form.html', {'medications': medications})
    
    return redirect('dashboard')


@login_required(login_url='login')
def pharmacy_stock_update(request, pk):
    if request.user.profile.role in ['Admin', 'PharmacyHead']:
        # Fetch the existing PharmacyStock entry by its ID
        pharmacy_stock = get_object_or_404(PharmacyStock, pk=pk)

        if request.method == 'POST':
            # Get updated values from the form
            medication_id = request.POST.get('medication')
            quantity = request.POST.get('quantity')
            unit = request.POST.get('unit')
            unit_price = request.POST.get('unit_price')
            transaction_type = request.POST.get('transaction_type')


            # Get the selected medication
            medication = get_object_or_404(Medication, id=medication_id)

            # Update the PharmacyStock entry
            pharmacy_stock.medication = medication
            
            pharmacy_stock.unit = unit
            pharmacy_stock.unit_price = unit_price
            pharmacy_stock.updated_by = request.user.profile
            

            if transaction_type == 'Addition':
                pharmacy_stock.quantity = int(quantity) + int(pharmacy_stock.quantity) 
                pharmacy_stock.save()
                PharmacyStockTransaction.objects.create(
                    pharmacy_stock = PharmacyStock.objects.get(id=pharmacy_stock.id),
                    transaction_type = 'Addition',
                    quantity = quantity,
                    created_by = request.user.profile,
                )


            if transaction_type == 'Removal':
                # Check if available quantity is greater than or equal to the quantity to be removed
                if int(pharmacy_stock.quantity) >= int(quantity):
                    pharmacy_stock.quantity -= int(quantity)  # Subtract the quantity
                    pharmacy_stock.save()  # Save the updated stock

                    # Create a new transaction record
                    PharmacyStockTransaction.objects.create(
                        pharmacy_stock=pharmacy_stock,
                        transaction_type='Removal',
                        quantity=quantity,
                        created_by=request.user.profile
                    )
                else:
                    # Handle the case where there is insufficient quantity
                    # You can raise an exception, return a response, or log a message
                    messages.success(request, "Insufficient stock available for removal.")

            if transaction_type == 'None':
                old_quantity = int(pharmacy_stock.quantity)
                new_quantity = int(quantity)
                delta = new_quantity - old_quantity
                pharmacy_stock.quantity = new_quantity
                pharmacy_stock.save()

                if delta > 0:
                    PharmacyStockTransaction.objects.create(
                        pharmacy_stock=pharmacy_stock,
                        transaction_type='Addition',
                        quantity=delta,
                        created_by=request.user.profile,
                    )
                elif delta < 0:
                    PharmacyStockTransaction.objects.create(
                        pharmacy_stock=pharmacy_stock,
                        transaction_type='Removal',
                        quantity=abs(delta),
                        created_by=request.user.profile,
                    )



            

            messages.success(request, 'Pharmacy stock updated successfully.')
            return redirect('manage_pharmacy_stock')

        medications = Medication.objects.all()
        return render(request, 'core/pharmacy_stock_update.html', {
            'pharmacy_stock': pharmacy_stock,
            'medications': medications,
        })
    
    return redirect('dashboard')

@login_required(login_url='login')
def pharmacy_stock_transaction(request):
    if request.user.profile.role in ['Admin']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        transactions = PharmacyStockTransaction.objects.filter(created_at__date__range=[start_date, end_date])

        return render(request, 'core/pharmacy_stock_transactions.html', {
            'transactions': transactions,
            'start_date': start_date,
            'end_date': end_date,
        })

    return redirect('dashboard')

@login_required(login_url='login')
def pharmacy_stock_delete(request, pk):
    if request.user.profile.role in ['Admin', 'PharmacyHead']:
        # Fetch the PharmacyStock entry by its ID
        pharmacy_stock = get_object_or_404(PharmacyStock, pk=pk)

        if request.method == 'POST' or request.method == 'GET':
            # Delete the PharmacyStock entry
            pharmacy_stock.delete()
            messages.success(request, 'Pharmacy stock deleted successfully.')
            return redirect('manage_pharmacy_stock')

    return redirect('dashboard')


# ====================== Bin Card ======================
# A bin card is a running ledger of every addition/removal for a medication,
# combining all of its batches into one chronological list with a running balance.

@login_required(login_url='login')
def bin_card_list(request):
    if request.user.profile.role not in ['Admin', 'PharmacyHead', 'Pharmacist', 'Accountant']:
        return redirect('dashboard')

    medications = Medication.objects.annotate(
        total_quantity=Sum('pharmacystock__quantity')
    ).order_by('name')

    return render(request, 'core/bin_card_list.html', {'medications': medications})


@login_required(login_url='login')
def bin_card_detail(request, medication_id):
    if request.user.profile.role not in ['Admin', 'PharmacyHead', 'Pharmacist', 'Accountant']:
        return redirect('dashboard')

    today = timezone.now().date()
    default_start = today - timedelta(days=6)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
    except ValueError:
        start_date = default_start
    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
    except ValueError:
        end_date = today

    medication = get_object_or_404(Medication, id=medication_id)
    stocks = PharmacyStock.objects.filter(medication=medication)

    # Build ALL chronological entries across every stock batch (same logic as before).
    # We need the full history to correctly compute the opening balance for the period.
    all_entries = []
    for stock in stocks:
        unit_price = stock.unit_price
        transactions = list(
            PharmacyStockTransaction.objects.filter(pharmacy_stock=stock)
            .select_related('created_by__user', 'patient')
            .order_by('created_at')
        )

        net_recorded = sum(
            txn.quantity if txn.transaction_type == 'Addition' else -txn.quantity
            for txn in transactions
        )
        opening = stock.quantity - net_recorded

        if opening != 0:
            qty_in = opening if opening > 0 else 0
            qty_out = abs(opening) if opening < 0 else 0
            all_entries.append({
                'date': stock.date_of_stock,
                'batch_number': stock.batch_number,
                'transaction_type': 'Opening Stock',
                'qty_in': qty_in,
                'qty_out': qty_out,
                'qty_in_value': qty_in * unit_price,
                'qty_out_value': qty_out * unit_price,
                'recorded_by': stock.created_by,
                'patient': None,
            })

        for txn in transactions:
            qty_in = txn.quantity if txn.transaction_type == 'Addition' else 0
            qty_out = txn.quantity if txn.transaction_type == 'Removal' else 0
            all_entries.append({
                'date': txn.created_at,
                'batch_number': stock.batch_number,
                'transaction_type': txn.transaction_type,
                'qty_in': qty_in,
                'qty_out': qty_out,
                'qty_in_value': qty_in * unit_price,
                'qty_out_value': qty_out * unit_price,
                'recorded_by': txn.created_by,
                'patient': txn.patient,
            })

    all_entries.sort(key=lambda e: e['date'])

    def entry_date(e):
        d = e['date']
        return d.date() if hasattr(d, 'date') else d

    # Compute opening balance = running total of everything BEFORE the period start.
    opening_bal_qty = 0
    opening_bal_value = Decimal('0')
    for e in all_entries:
        if entry_date(e) < start_date:
            opening_bal_qty += e['qty_in'] - e['qty_out']
            opening_bal_value += e['qty_in_value'] - e['qty_out_value']

    # Keep only entries within the selected period.
    period_entries = [e for e in all_entries if start_date <= entry_date(e) <= end_date]

    # First row is the computed opening balance for the period.
    entries = [{
        'date': start_date,
        'batch_number': '-',
        'transaction_type': 'Opening Balance',
        'qty_in': 0,
        'qty_out': 0,
        'qty_in_value': Decimal('0'),
        'qty_out_value': Decimal('0'),
        'recorded_by': None,
        'patient': None,
        'balance': opening_bal_qty,
        'balance_value': opening_bal_value,
    }]

    balance = opening_bal_qty
    balance_value = opening_bal_value
    for e in period_entries:
        balance += e['qty_in'] - e['qty_out']
        balance_value += e['qty_in_value'] - e['qty_out_value']
        e['balance'] = balance
        e['balance_value'] = balance_value
        entries.append(e)

    current_total = stocks.aggregate(total=Sum('quantity'))['total'] or 0
    current_stock_value = sum(s.quantity * s.unit_price for s in stocks)

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bin Card'

        headers = [
            'Date', 'Batch No.', 'Transaction',
            'Qty In', 'Qty In Value (₦)',
            'Qty Out', 'Qty Out Value (₦)',
            'Patient File No.',
            'Balance (Qty)', 'Balance Value (₦)',
            'Recorded By',
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        for row_num, entry in enumerate(entries, 2):
            date_val = entry['date']
            if hasattr(date_val, 'strftime'):
                date_val = date_val.strftime('%Y-%m-%d %H:%M')
            ws.cell(row=row_num, column=1, value=date_val)
            ws.cell(row=row_num, column=2, value=entry['batch_number'] or '-')
            ws.cell(row=row_num, column=3, value=entry['transaction_type'])
            ws.cell(row=row_num, column=4, value=entry['qty_in'] or 0)
            ws.cell(row=row_num, column=5, value=float(entry['qty_in_value']))
            ws.cell(row=row_num, column=6, value=entry['qty_out'] or 0)
            ws.cell(row=row_num, column=7, value=float(entry['qty_out_value']))
            ws.cell(row=row_num, column=8, value=entry['patient'].patient_file_number if entry['patient'] else '-')
            ws.cell(row=row_num, column=9, value=entry['balance'])
            ws.cell(row=row_num, column=10, value=float(entry['balance_value']))
            ws.cell(row=row_num, column=11, value=str(entry['recorded_by']) if entry['recorded_by'] else '-')

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename=bin_card_{medication.name}_{start_date}_to_{end_date}.xlsx'
        )
        wb.save(response)
        return response

    return render(request, 'core/bin_card_detail.html', {
        'medication': medication,
        'entries': entries,
        'current_total': current_total,
        'current_stock_value': current_stock_value,
        'start_date': start_date,
        'end_date': end_date,
    })


# View to manage Suppliers
@login_required(login_url='login')
def manage_suppliers(request):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Pharmacist':

        suppliers = Supplier.objects.all()
        if request.method == 'POST' and 'upload_file' in request.FILES:
            excel_file = request.FILES['upload_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)
            df = pd.read_excel(file_path)
            for index, row in df.iterrows():
                Supplier.objects.create(
                    name=row['name'],
                    contact_name=row['contact_name'],
                    contact_phone=row['contact_phone'],
                    contact_email=row['contact_email'],
                    address=row['address']
                )
            messages.success(request, 'Suppliers imported successfully.')
            return redirect('manage_suppliers')
        return render(request, 'core/manage_suppliers.html', {'suppliers': suppliers})

    return redirect('dashboard')

    # View to create or update a Supplier
@login_required(login_url='login')
def supplier_form(request, supplier_id=None):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'Pharmacist':

        if supplier_id:
            supplier = get_object_or_404(Supplier, id=supplier_id)
        else:
            supplier = None

        if request.method == 'POST':
            name = request.POST['name']
            contact_name = request.POST['contact_name']
            contact_phone = request.POST['contact_phone']
            contact_email = request.POST['contact_email']
            address = request.POST['address']
            if supplier:
                supplier.name = name
                supplier.contact_name = contact_name
                supplier.contact_phone = contact_phone
                supplier.contact_email = contact_email
                supplier.address = address
                supplier.save()
                messages.success(request, 'Supplier updated successfully.')
            else:
                Supplier.objects.create(
                    name=name,
                    contact_name=contact_name,
                    contact_phone=contact_phone,
                    contact_email=contact_email,
                    address=address
                )
                messages.success(request, 'Supplier created successfully.')
            return redirect('manage_suppliers')

        return render(request, 'core/supplier_form.html', {'supplier': supplier})
    
    return redirect('dashboard')



# ======================= lab ==============================


@login_required(login_url='login')
def labtech_dashboard(request):
    if request.user.profile.role in ['Admin', 'LabTech', 'Cashier']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        investigation_requests = InvestigationRequest.objects.filter(created_at__date__range=[start_date, end_date]).order_by('-created_at')

        # Grouping investigations by related invoices
        grouped_investigations = {}
        for investigation_request in investigation_requests:
            invoice_items = InvoiceItem.objects.filter(investigation_request=investigation_request, invoice__is_paid=True)
            for invoice_item in invoice_items:
                invoice_id = invoice_item.invoice.id
                if invoice_id not in grouped_investigations:
                    grouped_investigations[invoice_id] = {
                        'requests': [],
                        'status': investigation_request.status  # Assuming all requests in the same invoice have the same status
                    }
                grouped_investigations[invoice_id]['requests'].append(investigation_request)

        return render(request, 'core/labtech_dashboard.html', {
            'grouped_investigations': grouped_investigations,
            'start_date': start_date,
            'end_date': end_date,
        })

    return redirect('dashboard')


# @login_required(login_url='login')
# def labtech_dashboard(request):

#     if request.user.profile.role == 'Admin' or request.user.profile.role == 'LabTech' or request.user.profile.role == 'Cashier':
#         # Grouping investigations by related invoices
#         grouped_investigations = {}

#         # Fetching all requested investigations with paid invoices
#         investigation_requests = InvestigationRequest.objects.all().order_by('-created_at')

#         for investigation_request in investigation_requests:
#             # Filter for InvoiceItems related to the investigation and where the invoice is paid
#             invoice_items = InvoiceItem.objects.filter(investigation=investigation_request.investigation, invoice__is_paid=True)
            
#             for invoice_item in invoice_items:
#                 invoice_id = invoice_item.invoice.id
#                 if invoice_id not in grouped_investigations:
#                     grouped_investigations[invoice_id] = []
#                 grouped_investigations[invoice_id].append(investigation_request)

#         return render(request, 'core/labtech_dashboard.html', {'grouped_investigations': grouped_investigations})
    
#     return redirect('dashboard')



@login_required(login_url='login')
def create_grouped_investigation_results(request, invoice_id):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'LabTech' or request.user.profile.role == 'Cashier':
        # Fetch all investigation requests related to this invoice
        invoice_items = InvoiceItem.objects.filter(invoice__id=invoice_id, invoice__is_paid=True)
        investigation_requests = []
        for i in invoice_items:
            if i.investigation_request:
                investigation_requests.append(i.investigation_request)
        

        if request.method == 'POST':
            for investigation_request in investigation_requests:
                investigation = investigation_request.investigation
                result_type = investigation.result_type

                numeric_result = request.POST.get(f'numeric_result_{investigation_request.id}')
                reactive_non_reactive_result = request.POST.get(f'reactive_non_reactive_result_{investigation_request.id}')
                positive_negative_result = request.POST.get(f'positive_negative_result_{investigation_request.id}')
                text_result = request.POST.get(f'text_result_{investigation_request.id}')
                is_abnormal = request.POST.get(f'is_abnormal_{investigation_request.id}') == 'on'
                remarks = request.POST.get(f'remarks_{investigation_request.id}')

                # Creating the result based on the investigation's result type
                result_data = {
                    'request': investigation_request,
                    'result_type': result_type,
                    'recorded_by': request.user.profile,
                    'is_abnormal': is_abnormal,
                    'remarks': remarks,
                }
                
                if result_type == 'Numeric':
                    result_data['numeric_result'] = numeric_result
                elif result_type == 'Reactive/Non-Reactive':
                    result_data['text_result'] = reactive_non_reactive_result
                elif result_type == 'Positive/Negative':
                    result_data['text_result'] = positive_negative_result
                elif result_type == 'Text':
                    result_data['text_result'] = text_result
                
                InvestigationResult.objects.create(**result_data)
                investigation_request.status = 'Completed'
                investigation_request.save()
            
            messages.success(request, "All investigation results added successfully.")
            return redirect('labtech_dashboard')

        return render(request, 'core/create_grouped_investigation_results.html', {
            'investigation_requests': investigation_requests,
        })
    
    return redirect('dashboard')


from django.utils import timezone
from datetime import timedelta

@login_required(login_url='login')
def manage_investigation_requests(request):
    if request.user.profile.role in ['Admin', 'LabTech', 'Cashier']:
        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        pending_requests = InvestigationRequest.objects.filter(status='Requested', created_at__date__range=[start_date, end_date]).order_by('-created_at')

        return render(request, 'core/manage_investigation_requests.html', {
            'pending_requests': pending_requests,
            'start_date': start_date,
            'end_date': end_date,
        })

    return redirect('dashboard')


@login_required(login_url='login')
def delete_investigation_request(request, request_id):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'LabTech' or request.user.profile.role == 'Cashier':
        
        investigation_request = get_object_or_404(InvestigationRequest, id=request_id)
        investigation_request.delete()
        invoice_item = InvoiceItem.objects.get(investigation_request=investigation_request)
        invoice_item.delete()
        messages.success(request, "Investigation request deleted successfully.")
        return redirect('manage_investigation_requests')

    return redirect('dashboard')

@login_required(login_url='login')
def create_investigation_request(request):
    if request.user.profile.role in ['Admin', 'LabTech', 'Cashier']:
        patients = Patient.objects.all()
        investigations = Investigation.objects.all()

        if request.method == 'POST':
            patient_id = request.POST.get('patient_id')
            patient_name = request.POST.get('outpatient_name')
            patient_age = request.POST.get('outpatient_age')
            patient_gender = request.POST.get('outpatient_gender')
            investigation_ids = request.POST.getlist('investigation_id')  # Handle multiple investigations

            if patient_id:
                # If a patient is selected, get the patient object
                patient = Patient.objects.get(id=patient_id)
                patient_name = patient.first_name + ' ' + patient.last_name
                patient_age = None  # Assuming age is not needed if patient is in the system
                patient_gender = patient.gender
            elif patient_name and patient_age and patient_gender:
                # Outpatient case: Patient details are provided directly
                patient = None  # No patient object in this case
            else:
                messages.error(request, "Please select a patient or provide outpatient details.")
                return redirect('create_investigation_request')

            # Create an invoice for the patient or outpatient
            invoice = Invoice.objects.create(patient=patient, total_amount=0)

            total_amount = 0

            # Create investigation requests and add them to the invoice
            for investigation_id in investigation_ids:
                investigation = Investigation.objects.get(id=investigation_id)
                investigation_request = InvestigationRequest.objects.create(
                    patient=patient,  # For regular patients, this will be populated
                    patient_name=patient_name,  # Store outpatient or regular patient name
                    age=patient_age,  # Store outpatient or regular patient age
                    gender=patient_gender,  # Store outpatient or regular patient gender
                    investigation=investigation
                )
                # Add the investigation to the invoice
                invoice_item = InvoiceItem.objects.create(
                    invoice=invoice,
                    investigation_request=investigation_request,
                    quantity=1,  # Assuming each investigation is requested once
                    unit_price=investigation.price,
                    total_price=investigation.price
                )
                total_amount += invoice_item.total_price

            # Update the total amount of the invoice
            invoice.total_amount = total_amount
            invoice.save()

            messages.success(request, "Investigation request and invoice created successfully.")
            return redirect('manage_investigation_requests')

        return render(request, 'core/create_investigation_request.html', {
            'patients': patients,
            'investigations': investigations
        })

    return redirect('dashboard')



@login_required(login_url='login')
def manage_investigations(request):
    
    if request.user.profile.role == 'Admin' or request.user.profile.role == 'LabTech' or request.user.profile.role == 'Cashier':

        investigations = Investigation.objects.all()
        return render(request, 'core/manage_investigations.html', {'investigations': investigations})

    return redirect('dashboard')

@login_required(login_url='login')
def create_investigation(request):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'LabTech' or request.user.profile.role == 'Cashier':
        if request.method == 'POST':
            name = request.POST.get('name')
            unit_of_measurement = request.POST.get('unit_of_measurement')
            reference_range = request.POST.get('reference_range')
            result_type = request.POST.get('result_type')
            price = request.POST.get('price')

            Investigation.objects.create(
                name=name,
                unit_of_measurement=unit_of_measurement,
                reference_range=reference_range,
                result_type=result_type,
                price=price
            )
            messages.success(request, "Investigation created successfully.")
            return redirect('manage_investigations')

        return render(request, 'core/create_investigation.html')

    return redirect('dashboard')


@login_required(login_url='login')
def update_investigation(request, investigation_id):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'LabTech' or request.user.profile.role == 'Cashier':
        
        investigation = get_object_or_404(Investigation, id=investigation_id)

        if request.method == 'POST':
            investigation.name = request.POST.get('name')
            investigation.unit_of_measurement = request.POST.get('unit_of_measurement')
            investigation.reference_range = request.POST.get('reference_range')
            investigation.result_type = request.POST.get('result_type')
            investigation.price = request.POST.get('price')
            investigation.save()

            messages.success(request, "Investigation updated successfully.")
            return redirect('manage_investigations')

        return render(request, 'core/update_investigation.html', {'investigation': investigation})
    
    return redirect('dashboard')

@login_required(login_url='login')
def delete_investigation(request, investigation_id):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'LabTech' or request.user.profile.role == 'Cashier':
        
        investigation = get_object_or_404(Investigation, id=investigation_id)
        investigation.delete()
        messages.success(request, "Investigation deleted successfully.")
        return redirect('manage_investigations')
    
    return redirect('dashboard')


@login_required(login_url='login')
def view_grouped_investigation_results(request, invoice_id):

    if request.user.profile.role == 'Admin' or request.user.profile.role == 'LabTech' or request.user.profile.role == 'Cashier':
        # Fetch the invoice and related investigation requests
        invoice = get_object_or_404(Invoice, id=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        investigations_requests = []
        for i in invoice_items:
            if i.investigation_request:
                investigations_requests.append(i.investigation_request)
        

        context = {
            'invoice': invoice,
            'investigation_requests': investigations_requests,
        }

        return render(request, 'core/view_grouped_investigation_results.html', context)
    
    return redirect('dashboard')



@login_required(login_url='login')
def view_invoice(request, invoice_id):
    if request.user.profile.role in ['Admin', 'Cashier', 'Accountant']:
        # Get the invoice by ID
        invoice = get_object_or_404(Invoice, id=invoice_id)
        
        # Calculate total amount paid for this invoice
        total_paid = Payment.objects.filter(invoice=invoice).aggregate(total=Sum('amount'))['total'] or 0
        payments = Payment.objects.filter(invoice=invoice)
        first_payment = payments.first()
        created_by = first_payment.created_by if first_payment else None
        # Calculate total amount refunded for this invoice
        total_refunded = Refund.objects.filter(payment__invoice=invoice).aggregate(total=Sum('amount'))['total'] or 0
        
        # Calculate balance due
        balance_due = invoice.total_amount - total_paid + total_refunded
        
        # Retrieve all invoice items associated with this invoice
        items = invoice.items.all()
        
        # Build patient or custom name display
        if invoice.patient:
            patient_name = f"{invoice.patient.first_name} {invoice.patient.last_name} [ {invoice.patient.patient_file_number} ]"
        elif invoice.family:
            patient_name = f"{invoice.family.name} [ Family ]"
        elif invoice.custom_name:
            patient_name = invoice.custom_name
        else:
            patient_name = "N/A"
        
        # Prepare context data
        context = {
            'invoice': invoice,
            'total_paid': total_paid,
            'total_refunded': total_refunded,
            'balance_due': balance_due,
            'items': items,
            'patient_name': patient_name,
            'created_by': created_by,
            'payments': payments,
        }
        
        # Pass the invoice and other details to the template
        return render(request, 'core/view_invoice.html', context)

    return redirect('dashboard')



# ====================== Statistics Overview ======================
# A cross-department snapshot: patient volume, clinical activity, finance,
# pharmacy/inventory, and staff performance for a selected period. Each
# section links out to the existing detailed report for that area
# (Day Book, Bin Card, etc.) rather than duplicating their drill-down detail.

@login_required(login_url='login')
def statistics_dashboard(request):
    if request.user.profile.role != 'Admin':
        return redirect('dashboard')

    today = timezone.now().date()
    default_start = today - timedelta(days=6)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
    except ValueError:
        start_date = default_start
    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
    except ValueError:
        end_date = today

    def in_period(qs, field):
        return qs.filter(**{f'{field}__date__range': [start_date, end_date]})

    GENDER_LABELS = {'M': 'Male', 'F': 'Female', 'O': 'Other'}

    # ---------- 1. Patient ----------
    new_patients = in_period(Patient.objects.all(), 'created_at').count()
    total_active_patients = Patient.objects.filter(is_active=True).count()
    gender_breakdown = list(Patient.objects.values('gender').annotate(count=Count('id')).order_by('-count'))
    file_type_breakdown = list(
        Patient.objects.exclude(file_type__isnull=True).exclude(file_type='')
        .values('file_type').annotate(count=Count('id')).order_by('-count')
    )

    # ---------- 2. Clinical ----------
    appointment_status_breakdown = list(
        in_period(Appointment.objects.all(), 'appointment_date')
        .values('status').annotate(count=Count('id')).order_by('-count')
    )
    admissions_qs = in_period(Admission.objects.all(), 'admission_date')
    admissions_count = admissions_qs.count()
    avg_los = admissions_qs.filter(discharge_date__isnull=False) \
        .annotate(stay=F('discharge_date') - F('admission_date')) \
        .aggregate(avg=Avg('stay'))['avg']
    avg_length_of_stay_days = avg_los.days if avg_los else None

    top_diagnoses = list(
        in_period(Diagnosis.objects.all(), 'created_at')
        .values('name').annotate(count=Count('id')).order_by('-count')[:10]
    )
    investigation_status_breakdown = list(
        in_period(InvestigationRequest.objects.all(), 'date_requested')
        .values('status').annotate(count=Count('id')).order_by('-count')
    )

    # ---------- 3. Financial ----------
    period_payments = in_period(Payment.objects.all(), 'payment_date')
    period_refunds  = in_period(Refund.objects.all(), 'created_at')
    total_collected  = period_payments.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_refunded   = period_refunds.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    net_collected    = total_collected - total_refunded

    unpaid_total       = Invoice.objects.filter(is_paid=False).aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
    paid_on_unpaid     = Payment.objects.filter(invoice__is_paid=False).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    refunded_on_unpaid = Refund.objects.filter(payment__invoice__is_paid=False).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    outstanding_balance = unpaid_total - paid_on_unpaid + refunded_on_unpaid
    total_wallet_balance = Wallet.objects.aggregate(t=Sum('balance'))['t'] or Decimal('0')

    payment_method_breakdown = list(
        period_payments.values('payment_method').annotate(total=Sum('amount')).order_by('-total')
    )
    daily_collections = list(
        period_payments.annotate(day=TruncDate('payment_date'))
        .values('day').annotate(total=Sum('amount')).order_by('day')
    )

    # ---------- 4. Pharmacy ----------
    top_medications = list(
        in_period(InvoiceItem.objects.filter(medication__isnull=False), 'invoice__created_at')
        .values('medication__name').annotate(qty=Sum('quantity')).order_by('-qty')[:10]
    )
    stock_value = PharmacyStock.objects.aggregate(value=Sum(F('quantity') * F('unit_price')))['value'] or Decimal('0')
    out_of_stock_count = Medication.objects.annotate(
        total_quantity=Sum('pharmacystock__quantity')
    ).filter(Q(total_quantity__isnull=True) | Q(total_quantity__lte=0)).count()
    near_expiry_count = PharmacyStock.objects.filter(
        expiry_date__isnull=False,
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30),
        quantity__gt=0,
    ).count()

    # ---------- 5. Staff ----------
    top_doctors = list(
        in_period(Appointment.objects.all(), 'appointment_date')
        .values('doctor__profile__user__username').annotate(count=Count('id')).order_by('-count')[:10]
    )
    top_cashiers = list(
        period_payments.values('created_by__user__username')
        .annotate(total=Sum('amount')).order_by('-total')[:10]
    )

    # ---------- Excel export ----------
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()

        HDR_FILL = PatternFill('solid', fgColor='2a78d6')
        HDR_FONT = Font(bold=True, color='FFFFFF')

        def bold_row(ws, values):
            ws.append(values)
            for cell in ws[ws.max_row]:
                cell.font = HDR_FONT
                cell.fill = HDR_FILL

        def auto_width(ws, min_w=16, max_w=40):
            for col in ws.columns:
                best = min_w
                for cell in col:
                    if cell.value:
                        best = max(best, min(max_w, len(str(cell.value)) + 4))
                ws.column_dimensions[col[0].column_letter].width = best

        def bar_chart(ws, title, data_rows, labels_col, values_col, anchor, horizontal=True):
            if not data_rows:
                return
            n = len(data_rows)
            chart = BarChart()
            chart.type = 'bar' if horizontal else 'col'
            chart.title = title
            chart.style = 10
            chart.grouping = 'clustered'
            chart.width = 18
            chart.height = max(8, n * 1.2)
            chart.legend = None
            data_ref = Reference(ws, min_col=values_col, min_row=1, max_row=n + 1)
            cats_ref = Reference(ws, min_col=labels_col, min_row=2, max_row=n + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.series[0].graphicalProperties.solidFill = '2a78d6'
            ws.add_chart(chart, anchor)

        def pie_chart(ws, title, data_rows, labels_col, values_col, anchor):
            if not data_rows:
                return
            n = len(data_rows)
            chart = PieChart()
            chart.title = title
            chart.style = 10
            chart.width = 16
            chart.height = 14
            data_ref = Reference(ws, min_col=values_col, min_row=1, max_row=n + 1)
            cats_ref = Reference(ws, min_col=labels_col, min_row=2, max_row=n + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.dataLabels = openpyxl.chart.label.DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showCatName = True
            chart.dataLabels.showVal = False
            chart.set_categories(cats_ref)
            ws.add_chart(chart, anchor)

        def line_chart(ws, title, data_rows, labels_col, values_col, anchor):
            if not data_rows:
                return
            n = len(data_rows)
            chart = LineChart()
            chart.title = title
            chart.style = 10
            chart.width = 24
            chart.height = 12
            chart.legend = None
            data_ref = Reference(ws, min_col=values_col, min_row=1, max_row=n + 1)
            cats_ref = Reference(ws, min_col=labels_col, min_row=2, max_row=n + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.series[0].graphicalProperties.line.solidFill = '2a78d6'
            chart.series[0].graphicalProperties.line.width = 20000
            ws.add_chart(chart, anchor)

        # ---- Sheet 1: Summary ----
        ws1 = wb.active
        ws1.title = 'Summary'
        bold_row(ws1, ['Metric', 'Value'])
        for row in [
            ('Period', f'{start_date} to {end_date}'),
            ('New Patients (period)', new_patients),
            ('Total Active Patients', total_active_patients),
            ('Admissions (period)', admissions_count),
            ('Avg Length of Stay (days)', avg_length_of_stay_days or 'N/A'),
            ('Total Collected (₦)', float(total_collected)),
            ('Total Refunded (₦)', float(total_refunded)),
            ('Net Collected (₦)', float(net_collected)),
            ('Outstanding Balance (₦)', float(outstanding_balance)),
            ('Total Wallet Balances (₦)', float(total_wallet_balance)),
            ('Current Stock Value (₦)', float(stock_value)),
            ('Medications Out of Stock', out_of_stock_count),
            ('Batches Expiring in 30 Days', near_expiry_count),
        ]:
            ws1.append(list(row))
        auto_width(ws1)

        # ---- Sheet 2: Patients ----
        ws2 = wb.create_sheet('Patients')
        bold_row(ws2, ['Gender', 'Count'])
        for g in gender_breakdown:
            ws2.append([GENDER_LABELS.get(g['gender'], g['gender'] or 'Unknown'), g['count']])
        gender_end = ws2.max_row

        ws2.append([])
        ft_start = ws2.max_row + 1
        bold_row(ws2, ['File Type', 'Count'])
        for f in file_type_breakdown:
            ws2.append([f['file_type'], f['count']])
        auto_width(ws2)

        pie_chart(ws2, 'Gender Breakdown', gender_breakdown, 1, 2, 'D1')
        if file_type_breakdown:
            # Build a small helper: point chart references at the file-type block
            ft_n = len(file_type_breakdown)
            ft_chart = PieChart()
            ft_chart.title = 'File Type Breakdown'
            ft_chart.style = 10
            ft_chart.width = 16
            ft_chart.height = 14
            ft_data = Reference(ws2, min_col=2, min_row=ft_start, max_row=ft_start + ft_n)
            ft_cats = Reference(ws2, min_col=1, min_row=ft_start + 1, max_row=ft_start + ft_n)
            ft_chart.add_data(ft_data, titles_from_data=True)
            ft_chart.set_categories(ft_cats)
            ws2.add_chart(ft_chart, 'L1')

        # ---- Sheet 3: Clinical ----
        ws3 = wb.create_sheet('Clinical')
        bold_row(ws3, ['Appointment Status', 'Count'])
        for a in appointment_status_breakdown:
            ws3.append([a['status'], a['count']])
        appt_end = ws3.max_row

        ws3.append([])
        diag_start = ws3.max_row + 1
        bold_row(ws3, ['Top Diagnoses', 'Count'])
        for d in top_diagnoses:
            ws3.append([d['name'], d['count']])
        diag_end = ws3.max_row

        ws3.append([])
        inv_start = ws3.max_row + 1
        bold_row(ws3, ['Investigation Status', 'Count'])
        for i in investigation_status_breakdown:
            ws3.append([i['status'], i['count']])
        auto_width(ws3)

        bar_chart(ws3, 'Appointment Status', appointment_status_breakdown, 1, 2, 'D1')

        if top_diagnoses:
            dn = len(top_diagnoses)
            dc = BarChart()
            dc.type = 'bar'
            dc.title = 'Top Diagnoses'
            dc.style = 10
            dc.width = 18
            dc.height = max(8, dn * 1.2)
            dc.legend = None
            dc.add_data(Reference(ws3, min_col=2, min_row=diag_start, max_row=diag_end), titles_from_data=True)
            dc.set_categories(Reference(ws3, min_col=1, min_row=diag_start + 1, max_row=diag_end))
            dc.series[0].graphicalProperties.solidFill = 'e87ba4'
            ws3.add_chart(dc, 'D22')

        if investigation_status_breakdown:
            inv_n = len(investigation_status_breakdown)
            ic = BarChart()
            ic.type = 'bar'
            ic.title = 'Investigation Status'
            ic.style = 10
            ic.width = 18
            ic.height = max(8, inv_n * 1.2)
            ic.legend = None
            ic.add_data(Reference(ws3, min_col=2, min_row=inv_start, max_row=inv_start + inv_n), titles_from_data=True)
            ic.set_categories(Reference(ws3, min_col=1, min_row=inv_start + 1, max_row=inv_start + inv_n))
            ic.series[0].graphicalProperties.solidFill = 'eda100'
            ws3.add_chart(ic, 'L22')

        # ---- Sheet 4: Financial ----
        ws4 = wb.create_sheet('Financial')
        bold_row(ws4, ['Date', 'Collections (₦)'])
        for d in daily_collections:
            ws4.append([str(d['day']), float(d['total'])])
        daily_end = ws4.max_row

        ws4.append([])
        pm_start = ws4.max_row + 1
        bold_row(ws4, ['Payment Method', 'Total (₦)'])
        for p in payment_method_breakdown:
            ws4.append([p['payment_method'] or 'Unknown', float(p['total'])])
        pm_end = ws4.max_row
        auto_width(ws4)

        line_chart(ws4, f'Daily Collections {start_date} to {end_date}', daily_collections, 1, 2, 'D1')

        if payment_method_breakdown:
            pm_n = len(payment_method_breakdown)
            pc = PieChart()
            pc.title = 'Payment Methods'
            pc.style = 10
            pc.width = 16
            pc.height = 14
            pc.add_data(Reference(ws4, min_col=2, min_row=pm_start, max_row=pm_end), titles_from_data=True)
            pc.set_categories(Reference(ws4, min_col=1, min_row=pm_start + 1, max_row=pm_end))
            ws4.add_chart(pc, 'D22')

        # ---- Sheet 5: Pharmacy ----
        ws5 = wb.create_sheet('Pharmacy')
        bold_row(ws5, ['Medication', 'Qty Sold'])
        for m in top_medications:
            ws5.append([m['medication__name'], int(m['qty'])])
        auto_width(ws5)
        bar_chart(ws5, 'Top Medications Dispensed', top_medications, 1, 2, 'D1')

        # ---- Sheet 6: Staff ----
        ws6 = wb.create_sheet('Staff')
        bold_row(ws6, ['Doctor', 'Appointments'])
        for d in top_doctors:
            ws6.append([d['doctor__profile__user__username'] or 'Unassigned', d['count']])
        doc_end = ws6.max_row

        ws6.append([])
        cash_start = ws6.max_row + 1
        bold_row(ws6, ['Cashier', 'Collections (₦)'])
        for c in top_cashiers:
            ws6.append([c['created_by__user__username'] or 'Unknown', float(c['total'])])
        cash_end = ws6.max_row
        auto_width(ws6)

        bar_chart(ws6, 'Top Doctors by Appointments', top_doctors, 1, 2, 'D1')

        if top_cashiers:
            cash_n = len(top_cashiers)
            cc = BarChart()
            cc.type = 'col'
            cc.title = 'Top Cashiers by Collections (₦)'
            cc.style = 10
            cc.width = 18
            cc.height = 12
            cc.legend = None
            cc.add_data(Reference(ws6, min_col=2, min_row=cash_start, max_row=cash_end), titles_from_data=True)
            cc.set_categories(Reference(ws6, min_col=1, min_row=cash_start + 1, max_row=cash_end))
            cc.series[0].graphicalProperties.solidFill = '1baf7a'
            ws6.add_chart(cc, 'D22')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=statistics_{start_date}_to_{end_date}.xlsx'
        wb.save(response)
        return response

    # ---------- JSON for charts ----------
    chart_data = json.dumps({
        'gender': [{'label': GENDER_LABELS.get(g['gender'], g['gender'] or 'Unknown'), 'count': g['count']} for g in gender_breakdown],
        'file_type': [{'label': f['file_type'], 'count': f['count']} for f in file_type_breakdown],
        'appt_status': [{'label': a['status'], 'count': a['count']} for a in appointment_status_breakdown],
        'top_diagnoses': [{'label': d['name'], 'count': d['count']} for d in top_diagnoses],
        'inv_status': [{'label': i['status'], 'count': i['count']} for i in investigation_status_breakdown],
        'top_meds': [{'label': m['medication__name'], 'qty': int(m['qty'])} for m in top_medications],
        'top_doctors': [{'label': d['doctor__profile__user__username'] or 'Unassigned', 'count': d['count']} for d in top_doctors],
        'top_cashiers': [{'label': c['created_by__user__username'] or 'Unknown', 'total': float(c['total'])} for c in top_cashiers],
        'payment_methods': [{'label': p['payment_method'] or 'Unknown', 'total': float(p['total'])} for p in payment_method_breakdown],
        'daily': [{'date': str(d['day']), 'total': float(d['total'])} for d in daily_collections],
    })

    return render(request, 'core/statistics_dashboard.html', {
        'start_date': start_date,
        'end_date': end_date,
        'new_patients': new_patients,
        'total_active_patients': total_active_patients,
        'admissions_count': admissions_count,
        'avg_length_of_stay_days': avg_length_of_stay_days,
        'total_collected': total_collected,
        'total_refunded': total_refunded,
        'net_collected': net_collected,
        'outstanding_balance': outstanding_balance,
        'total_wallet_balance': total_wallet_balance,
        'stock_value': stock_value,
        'out_of_stock_count': out_of_stock_count,
        'near_expiry_count': near_expiry_count,
        'chart_data': chart_data,
    })


# View to manage Medications
@login_required(login_url='login')
def manage_user_profiles(request):

    if request.user.profile.role == 'Admin':

        profiles = Profile.objects.all()
        return render(request, 'core/manage_user_profiles.html', {'profiles': profiles})
    
    return redirect('dashboard')


# delete profile
@login_required(login_url='login')
def delete_user_profiles(request, profile_id):

    if request.user.profile.role == 'Admin':

        profile = Profile.objects.get(id=profile_id)
        profile.delete()
        messages.success(request, "User Profile deleted successfully.")
        return redirect('manage_user_profiles')
    
    return redirect('dashboard')








@login_required(login_url='login')
@transaction.atomic
def discharge_patient(request, admission_id):
    if request.user.profile.role in ['Admin', 'Doctor']:
        try:
            admission = Admission.objects.get(id=admission_id)
            
            # Get the current time as a timezone-aware datetime
            discharge_date = timezone.now()
            
            # Ensure both dates are timezone-aware
            if timezone.is_naive(admission.admission_date):
                admission.admission_date = timezone.make_aware(admission.admission_date, timezone.get_default_timezone())
            
            # Calculate the total number of days for the admission
            admission_duration = (discharge_date - admission.admission_date).days + 1  # +1 to include the discharge day
            
            # Get the admission service cost
            admission_service = Service.objects.get(name='Admission')
            total_cost = admission_service.cost * admission_duration
            
            # Create Invoice
            invoice = Invoice.objects.create(patient=admission.patient, total_amount=total_cost)
            
            # Create InvoiceItem for the admission
            InvoiceItem.objects.create(
                invoice=invoice,
                service=admission_service,
                quantity=admission_duration,
                unit_price=admission_service.cost,
                total_price=total_cost
            )
            
            # Update Admission status
            admission.status = 'Completed'
            admission.discharge_date = discharge_date
            admission.save()

            messages.success(request, "Patient discharged and invoice created.")
            return redirect('admissions')
        
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('admissions')

    return redirect('dashboard')


@login_required(login_url='login')
def create_diagnosis(request, appointment_id=None, admission_id=None):
    if request.user.profile.role in ['Admin', 'Doctor']:
        if request.method == 'POST':
            name = request.POST.get('name')
            if not name:
                return render(request, 'core/create_diagnosis.html', {'error': 'Diagnosis name is required'})
            
            # Get current user as doctor
            doctor = request.user
            
            if appointment_id:
                appointment = get_object_or_404(Appointment, id=appointment_id)
                diagnosis = Diagnosis.objects.create(
                    appointment=appointment,
                    doctor=doctor,
                    name=name
                )
                return redirect('doctor_dashboard')
            
            if admission_id:
                admission = get_object_or_404(Admission, id=admission_id)
                diagnosis = Diagnosis.objects.create(
                    admission=admission,
                    doctor=doctor,
                    name=name
                )
                return redirect('admissions')

        context = {
            'appointment_id': appointment_id,
            'admission_id': admission_id,
        }

        return render(request, 'core/create_diagnosis.html', context)

    return redirect('dashboard')





@login_required(login_url='login')
def dispense_medication(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    invoice_items = InvoiceItem.objects.filter(invoice=invoice)

    unavailable_items = []  # To store medications that are out of stock

    # Check stock availability for each medication
    for item in invoice_items:
        medication_stock = PharmacyStock.objects.filter(medication=item.medication).first()

        if not medication_stock or medication_stock.quantity < item.quantity:
            available_quantity = medication_stock.quantity if medication_stock else 0
            unavailable_items.append(f"{item.medication.name}: Available - {available_quantity}, Required - {item.quantity}")

    # If any medication is unavailable, return an error message
    if unavailable_items:
        messages.error(request, "Medication cannot be dispensed due to insufficient stock:\n" + "\n".join(unavailable_items))
        return redirect('manage_medication_invoices')

    # Deduct stock and create stock transaction
    for item in invoice_items:
        medication_stock = PharmacyStock.objects.get(medication=item.medication)
        medication_stock.quantity -= item.quantity
        medication_stock.save()

        # Create a stock transaction record
        PharmacyStockTransaction.objects.create(
            pharmacy_stock=medication_stock,
            transaction_type='Removal',
            quantity=item.quantity,
            patient=invoice.patient,
            created_by=request.user.profile
        )

    # Update invoice status
    invoice.status = 'Dispensed'
    invoice.save()

    messages.success(request, f"Medication for Invoice #{invoice.invoice_number} has been successfully dispensed.")
    return redirect('manage_medication_invoices')









@login_required(login_url='login')
def payments(request):
    if request.user.profile.role in ['Admin', 'Cashier', 'Accountant']:
        payments = Payment.objects.select_related(
            'invoice__patient', 'invoice__family', 'created_by__user'
        ).order_by('-payment_date')

        # Initialize total amount
        total_amount = None
        filters_applied = False

        # Filter by created_by if provided
        created_by = request.GET.get('created_by')
        if created_by:
            payments = payments.filter(created_by=Profile.objects.get(id=created_by))
            filters_applied = True

        # Filter by date range if provided
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if not start_date and not end_date and not filters_applied:
            start_date = timezone.now().date()
            end_date = start_date

        if start_date and end_date:
            payments = payments.filter(payment_date__date__range=[start_date, end_date])
            filters_applied = True

        # Initialize `invoice_items` to aggregate total for specific services
        invoice_items = InvoiceItem.objects.none()
        service_type = request.GET.get('service_type')

        # Apply service_type filter to isolate relevant `InvoiceItem`s
        if service_type:
            if service_type == 'Medication':
                # Filter payments by Medication and get only relevant InvoiceItems
                payments = payments.filter(invoice__items__medication__isnull=False).distinct()
                invoice_items = InvoiceItem.objects.filter(invoice__payment__in=payments, medication__isnull=False)
            elif service_type == 'Investigation':
                payments = payments.filter(invoice__items__investigation_request__isnull=False).distinct()
                invoice_items = InvoiceItem.objects.filter(invoice__payment__in=payments, investigation_request__isnull=False)
            elif service_type == 'Others':
                payments = payments.filter(invoice__items__medication__isnull=True, invoice__items__investigation_request__isnull=False, invoice__items__service__isnull=True).distinct()
                invoice_items = InvoiceItem.objects.filter(invoice__payment__in=payments, descriptions__isnull=False)
            else:
                payments = payments.filter(invoice__items__service__id=service_type).distinct()
                invoice_items = InvoiceItem.objects.filter(invoice__payment__in=payments, service__id=service_type)

            filters_applied = True

        # Filter by payment method if provided
        payment_method = request.GET.get('payment_method')
        if payment_method:
            payments = payments.filter(payment_method=payment_method)
            filters_applied = True

        # Calculate total amount based on service_type-filtered invoice items only
        if filters_applied and service_type:
            total_amount = invoice_items.aggregate(
                total=Sum(F('quantity') * F('unit_price'))
            )['total'] or 0
        else:
            total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0

        # Export: stream directly to Excel without building an intermediate list.
        # prefetch_related pulls items in one query instead of one per payment.
        if 'export' in request.GET:
            export_qs = payments.prefetch_related(
                'invoice__items__service',
                'invoice__items__medication',
                'invoice__items__investigation_request__investigation',
            )
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet('Payments')
            ws.append(['Payment No.', 'Amount', 'Payment Method', 'Date', 'Created By', 'Patient/Family', 'Items'])
            for p in export_qs:
                inv = p.invoice
                if inv.patient:
                    detail = f"{inv.patient.first_name} {inv.patient.last_name} [{inv.patient.patient_file_number}]"
                elif inv.family:
                    detail = f"{inv.family.name} [Family]"
                elif inv.custom_name:
                    detail = inv.custom_name
                else:
                    detail = 'N/A'
                parts = []
                for item in inv.items.all():
                    if item.medication_id:
                        parts.append(f"Medication: {item.medication.name} x {item.quantity}")
                    elif item.investigation_request_id:
                        parts.append(f"Investigation: {item.investigation_request.investigation.name} x {item.quantity}")
                    elif item.service_id:
                        parts.append(f"Service: {item.service.name} x {item.quantity}")
                    elif item.descriptions:
                        parts.append(f"Other: {item.descriptions} x {item.quantity}")
                ws.append([
                    p.payment_number or str(p.id),
                    float(p.amount),
                    p.payment_method,
                    p.payment_date.strftime('%Y-%m-%d'),
                    p.created_by.user.username if p.created_by else '-',
                    detail,
                    '; '.join(parts),
                ])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=payments_{start_date}_to_{end_date}.xlsx'
            wb.save(response)
            return response

        payment_data = []
        for payment in payments:
            invoice = payment.invoice
            try:
                if invoice.patient:
                    detail = f"{invoice.patient.first_name} {invoice.patient.last_name} [File Number: {invoice.patient.patient_file_number}]"
                elif invoice.family:
                    detail = f"{invoice.family.name} [Family]"
                elif invoice.custom_name:
                    detail = f"{invoice.custom_name} [Custom Patient]"
                else:
                    detail = "N/A"
            except Invoice.DoesNotExist:
                detail = "Invoice Not Found"

            payment_data.append({
                'payment': payment,
                'details': detail
            })

        return render(request, 'core/payments.html', {
            'profiles': Profile.objects.filter(role='Cashier').all(),
            'payments': payment_data,
            'service_types': Service.objects.all(),
            'total_amount': total_amount,
            'payment_methods': Payment.objects.values_list('payment_method', flat=True).distinct(),
            'default_start_date': start_date,
            'default_end_date': end_date,
        })

    return redirect('dashboard')



@login_required(login_url='login')
def edit_payment(request, payment_id):
    if request.user.profile.role not in ['Admin', 'Accountant']:
        return redirect('dashboard')

    payment = Payment.objects.get(id=payment_id)
    if request.method == 'POST':
        amount = request.POST.get('amount')
        payment_method = request.POST.get('payment_method')
        payment_date = request.POST.get('payment_date')

        payment.amount = amount
        payment.payment_method = payment_method
        payment.payment_date = payment_date
        payment.save()
        messages.success(request, f"Payment Updated Succesfully")
        return redirect('payments')


    return render(request, 'core/edit_payment.html', {'payment': payment})


@login_required(login_url='login')
def delete_payment(request, payment_id):
    if request.user.profile.role not in ['Admin', 'Accountant']:
        return redirect('dashboard')

    payment = Payment.objects.get(id=payment_id)
    payment.delete()

    return redirect('payments')


# ====================== Day Book ======================
# Each row = one cashier × one date.
# Columns: CASHIER, DATE, AMOUNT PAID,
#          [payment methods from DB],
#          [item categories from DB: Service.name / "Medication" / "Lab" / item.descriptions]
# Amounts for item categories are proportionally allocated from each payment
# across the invoice items (payment.amount × item.total_price / invoice.total_amount).

@login_required(login_url='login')
def day_book(request):
    role = request.user.profile.role
    if role not in ['Admin', 'Accountant', 'Cashier']:
        return redirect('dashboard')

    today = timezone.now().date()
    default_start = today - timedelta(days=6)

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
    except ValueError:
        start_date = default_start

    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
    except ValueError:
        end_date = today

    payments = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date]
    ).select_related(
        'invoice', 'created_by__user'
    ).prefetch_related(
        'invoice__items__service',
        'invoice__items__medication',
        'invoice__items__investigation_request__investigation',
    ).order_by('payment_date')

    # Cashiers only ever see their own day book rows, never other cashiers'.
    if role == 'Cashier':
        payments = payments.filter(created_by=request.user.profile)

    def item_category(item):
        if item.service_id:
            return item.service.name
        if item.medication_id:
            return 'Medication'
        if item.investigation_request_id:
            return 'Lab'
        return item.descriptions or 'Other'

    rows = {}
    payment_methods_seen = set()
    categories_seen = set()

    for payment in payments:
        cashier = (
            payment.created_by.user.username
            if payment.created_by and payment.created_by.user
            else 'Unknown'
        )
        date = payment.payment_date.date()
        key = (cashier, date)

        if key not in rows:
            rows[key] = {
                'cashier': cashier,
                'date': date,
                'amount_paid': Decimal('0'),
                'payment_methods': {},
                'categories': {},
            }

        row = rows[key]
        row['amount_paid'] += payment.amount

        method = payment.payment_method or 'Unknown'
        payment_methods_seen.add(method)
        row['payment_methods'][method] = row['payment_methods'].get(method, Decimal('0')) + payment.amount

        invoice = payment.invoice
        if invoice:
            items = list(invoice.items.all())
            invoice_total = invoice.total_amount or Decimal('0')
            if items and invoice_total > 0:
                for item in items:
                    share = item.total_price / invoice_total
                    allocated = payment.amount * share
                    cat = item_category(item)
                    categories_seen.add(cat)
                    row['categories'][cat] = row['categories'].get(cat, Decimal('0')) + allocated

    payment_method_cols = sorted(payment_methods_seen)
    category_cols = sorted(categories_seen)

    sorted_rows = sorted(rows.values(), key=lambda r: (r['date'], r['cashier']))
    for row in sorted_rows:
        row['method_values'] = [row['payment_methods'].get(m, Decimal('0')) for m in payment_method_cols]
        row['category_values'] = [row['categories'].get(c, Decimal('0')) for c in category_cols]

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Day Book'

        headers = ['CASHIER', 'DATE', 'AMOUNT PAID'] + payment_method_cols + category_cols
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header.upper())
            cell.font = Font(bold=True)

        for row_num, row in enumerate(sorted_rows, 2):
            values = (
                [row['cashier'], row['date'].strftime('%Y-%m-%d'), float(row['amount_paid'])]
                + [float(v) for v in row['method_values']]
                + [float(v) for v in row['category_values']]
            )
            for col_num, val in enumerate(values, 1):
                ws.cell(row=row_num, column=col_num, value=val)

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename=day_book_{start_date}_to_{end_date}.xlsx'
        )
        wb.save(response)
        return response

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'rows': sorted_rows,
        'payment_method_cols': payment_method_cols,
        'category_cols': category_cols,
    }

    return render(request, 'core/day_book.html', context)


@login_required(login_url='login')
def create_invoice_and_items(request):
    if request.user.profile.role in ['Admin', 'Cashier']:
        patients = Patient.objects.all()
        families = Family.objects.all()
        services = Service.objects.all()
        medications = Medication.objects.all()
        investigation_requests = InvestigationRequest.objects.all()

        if request.method == 'POST':
            patient_id = request.POST.get('patient_id')
            family_id = request.POST.get('family_id')
            custom_name = request.POST.get('custom_name')

            if patient_id:
                patient = get_object_or_404(Patient, id=patient_id)
                invoice = Invoice.objects.create(patient=patient, total_amount=0)
            elif family_id:
                family = get_object_or_404(Family, id=family_id)
                invoice = Invoice.objects.create(family=family, total_amount=0)
            elif custom_name:
                invoice = Invoice.objects.create(custom_name=custom_name, total_amount=0)
            else:
                messages.error(request, "Please select a patient, family, or enter a name.")
                return redirect('create_invoice_and_items')

            total_amount = 0

            item_types = request.POST.getlist('item_type')
            service_ids = request.POST.getlist('service_id')
            medication_ids = request.POST.getlist('medication_id')
            investigationrequest_ids = request.POST.getlist('investigationrequest_id')
            other_items = request.POST.getlist('other_item')
            unit_prices = request.POST.getlist('unit_price')
            quantities = request.POST.getlist('quantity')

            for i in range(len(item_types)):
                item_type = item_types[i]
                quantity = int(quantities[i])
                unit_price = float(unit_prices[i])
                total_price = quantity * unit_price

                if item_type == 'Service':
                    service = get_object_or_404(Service, id=service_ids[i])
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        service=service,
                        quantity=quantity,
                        unit_price=unit_price,
                        total_price=total_price
                    )
                elif item_type == 'Medication':
                    medication = get_object_or_404(Medication, id=medication_ids[i])
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        medication=medication,
                        quantity=quantity,
                        unit_price=unit_price,
                        total_price=total_price
                    )
                elif item_type == 'InvestigationRequest':
                    investigation_request = get_object_or_404(InvestigationRequest, id=investigationrequest_ids[i])
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        investigation_request=investigation_request,
                        quantity=quantity,
                        unit_price=unit_price,
                        total_price=total_price
                    )
                elif item_type == 'Others':
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        quantity=quantity,
                        unit_price=unit_price,
                        total_price=total_price,
                        descriptions=other_items[i]
                    )

                total_amount += total_price

            invoice.total_amount = total_amount
            invoice.save()

            messages.success(request, "Invoice created successfully.")
            return redirect('cashier_dashboard')  # Redirect to your invoices list or any other page

        context = {
            'patients': patients,
            'families': families,
            'services': services,
            'medications': medications,
            'investigation_requests': investigation_requests,
        }
        return render(request, 'core/create_invoice_and_items.html', context)
    
    return redirect('dashboard')


@login_required(login_url='login')
def wallet_list(request):
    if request.user.profile.role in ['Admin', 'Cashier', 'Accountant']:
        wallets = Wallet.objects.all()

        context = {
            'wallets': wallets,
        }
        return render(request, 'core/wallet_list.html', context)
    return redirect('dashboard')


@login_required(login_url='login')
def transfer_funds(request):
    if request.user.profile.role in ['Admin', 'Cashier', 'Accountant']:
        if request.method == 'POST':
            source_wallet_id = request.POST.get('source_wallet')
            target_wallet_id = request.POST.get('target_wallet')
            amount = Decimal(request.POST.get('amount'))

            source_wallet = get_object_or_404(Wallet, id=source_wallet_id)
            target_wallet = get_object_or_404(Wallet, id=target_wallet_id)

            # Check if the source wallet has enough balance
            if amount > 0:
                # Deduct from source wallet
                if source_wallet.balance >= amount:
                    # Debit the source wallet
                    Transaction.objects.create(
                        wallet=source_wallet,
                        amount=amount,
                        transaction_type='Debit',
                        description=f'Transfer to {target_wallet}',
                        created_by=request.user.profile
                    )
                    source_wallet.balance -= amount
                    source_wallet.save()

                    # Credit the target wallet
                    Transaction.objects.create(
                        wallet=target_wallet,
                        amount=amount,
                        transaction_type='Transfer',
                        description=f'Transfer from {source_wallet}',
                        created_by=request.user.profile
                    )
                    target_wallet.balance += amount
                    target_wallet.save()

                    messages.success(request, 'Funds transferred successfully.')
                    return redirect('wallet_list')
                else:
                    messages.error(request, 'Insufficient balance in the source wallet.')

        wallets = Wallet.objects.all()  # List all wallets for both source and target

        context = {
            'wallets': wallets,
        }
        return render(request, 'core/transfer_funds.html', context)
    return redirect('dashboard')

@login_required(login_url='login')
def view_wallet_transactions(request, wallet_id):
    if request.user.profile.role in ['Admin', 'Cashier', 'Accountant']:
        wallet = get_object_or_404(Wallet, id=wallet_id)

        today = timezone.now().date()
        default_start = today - timedelta(days=6)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        except ValueError:
            start_date = default_start
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except ValueError:
            end_date = today

        transactions = wallet.transaction_set.filter(date__date__range=[start_date, end_date]).order_by('-date')

        context = {
            'wallet': wallet,
            'transactions': transactions,
            'start_date': start_date,
            'end_date': end_date,
        }
        return render(request, 'core/view_wallet_transactions.html', context)
    return redirect('dashboard')





@login_required(login_url='login')
def credit_wallet_view(request):
    if request.user.profile.role in ['Admin', 'Cashier', 'Accountant']:
        if request.method == 'POST':
            wallet_id = request.POST.get('wallet')
            amount_str = request.POST.get('amount')
            description = request.POST.get('description', 'Wallet Credit')
            transaction_method = request.POST.get('transaction_method')

            if not transaction_method:
                messages.error(request, "Please select a transaction method.")
                return redirect('credit_wallet')

            try:
                amount = Decimal(amount_str)
            except ValueError:
                messages.error(request, "Invalid amount.")
                return redirect('credit_wallet')

            wallet = get_object_or_404(Wallet, id=wallet_id)

            # Credit the wallet
            wallet.balance += amount
            wallet.save()

            # Create a transaction record with the transaction method
            transaction_object = Transaction.objects.create(
                wallet=wallet,
                amount=amount,
                transaction_type='Credit',
                transaction_method=transaction_method,  # Capture the transaction method
                description=description,
                created_by=request.user.profile
            )

            messages.success(request, f"Wallet for {wallet} credited successfully with ₦{amount} via {transaction_method}.")
            return redirect('wallet_funding_receipt', transaction_id=transaction_object.id)

        wallets = Wallet.objects.all()
        return render(request, 'core/credit_wallet.html', {'wallets': wallets})

    return redirect('dashboard')


@login_required(login_url='login')
def wallet_funding_receipt(request, transaction_id):
    # Fetch the transaction details
    transaction = get_object_or_404(Transaction, id=transaction_id, transaction_type='Credit')

    # Ensure that the transaction is for wallet funding
    if transaction.transaction_type != 'Credit':
        return redirect('cashier_dashboard')

    # Render the receipt template with transaction details
    context = {
        'transaction': transaction
    }

    return render(request, 'core/wallet_funding_receipt.html', context)


@login_required(login_url='login')
def view_paid_invoice(request, invoice_id):
    if request.user.profile.role not in ['Admin', 'Accountant']:
        return redirect('dashboard')

    invoice = get_object_or_404(Invoice, id=invoice_id, is_paid=True)

    if request.method == 'POST':
        item_ids_to_remove = request.POST.getlist('item_ids')
        quantities = request.POST.getlist('quantities')
        items_to_refund = []
        
        # Loop through items to update quantities and find those to remove
        for index, item in enumerate(invoice.items.all()):
            new_quantity = int(quantities[index])
            
            # If the quantity has changed, update the item quantity and total price
            if item.quantity != new_quantity:
                # Calculate the price difference and adjust the total price accordingly
                difference_in_total = (new_quantity - item.quantity) * item.unit_price
                item.quantity = new_quantity
                item.total_price = item.unit_price * new_quantity
                item.save()
                
                # Update the invoice total
                invoice.total_amount += difference_in_total
                invoice.save()
                
            # Check if this item is selected for removal
            if str(item.id) in item_ids_to_remove:
                items_to_refund.append(item)
                item.delete()  # Remove the item from the invoice
        
        # Update the invoice total after any removals
        new_total = sum(item.total_price for item in invoice.items.all())
        invoice.total_amount = new_total
        invoice.save()

        # Process refunds for removed items
        for item in items_to_refund:
            refund = Refund.objects.create(
                payment=Payment.objects.filter(invoice=invoice).first(),
                amount=item.total_price,
                created_by=request.user.profile
            )

            # Update the wallet
            wallet = Wallet.objects.filter(patient=invoice.patient).first() or Wallet.objects.filter(family=invoice.family).first()
            if wallet:
                Transaction.objects.create(
                    wallet=wallet,
                    amount=item.total_price,
                    transaction_type='Credit',
                    transaction_method='Refund',
                    description=f"Refund for item {item.id} from Invoice #{invoice.invoice_number}",
                    created_by=request.user.profile
                )

                wallet.balance += item.total_price
                wallet.save()

        messages.success(request, "Selected items have been removed, quantities updated, and refunds processed.")
        return redirect('view_invoice', invoice_id)

    invoice_items = invoice.items.all()
    context = {
        'invoice': invoice,
        'invoice_items': invoice_items,
    }
    return render(request, 'core/view_paid_invoice.html', context)


# Inventory Category Views
@login_required(login_url='login')
def category_list(request):
    categories = InventoryCategory.objects.all()
    return render(request, 'core/category_list.html', {'categories': categories})

@login_required(login_url='login')
def category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')

        InventoryCategory.objects.create(name=name, description=description)
        messages.success(request, "Category created successfully.")
        return redirect('category_list')

    return render(request, 'core/category_form.html')

@login_required(login_url='login')
def category_update(request, pk):
    category = get_object_or_404(InventoryCategory, pk=pk)

    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.description = request.POST.get('description')
        category.save()
        messages.success(request, "Category updated successfully.")
        return redirect('category_list')

    return render(request, 'core/category_form.html', {'category': category})

@login_required(login_url='login')
def category_delete(request, pk):
    category = get_object_or_404(InventoryCategory, pk=pk)
    category.delete()
    messages.success(request, "Category deleted successfully.")
    return redirect('category_list')

# Inventory Item Views
@login_required(login_url='login')
def item_list(request):
    items = InventoryItem.objects.all()
    return render(request, 'core/item_list.html', {'items': items})

@login_required(login_url='login')
def item_create(request):
    categories = InventoryCategory.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        quantity = request.POST.get('quantity')
        unit_price = request.POST.get('unit_price')
        reorder_level = request.POST.get('reorder_level')

        InventoryItem.objects.create(
            name=name,
            category_id=category_id,
            quantity=quantity,
            unit_price=unit_price,
            reorder_level=reorder_level
        )
        messages.success(request, "Item created successfully.")
        return redirect('item_list')

    return render(request, 'core/item_form.html', {'categories': categories})

@login_required(login_url='login')
def item_update(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    categories = InventoryCategory.objects.all()

    if request.method == 'POST':
        item.name = request.POST.get('name')
        item.category_id = request.POST.get('category')
        item.quantity = request.POST.get('quantity')
        item.unit_price = request.POST.get('unit_price')
        item.reorder_level = request.POST.get('reorder_level')
        item.save()
        messages.success(request, "Item updated successfully.")
        return redirect('item_list')

    return render(request, 'core/item_form.html', {'item': item, 'categories': categories})

@login_required(login_url='login')
def item_delete(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    item.delete()
    messages.success(request, "Item deleted successfully.")
    return redirect('item_list')


@login_required(login_url='login')
def transaction_list(request):
    transactions = InventoryTransaction.objects.all()
    items = InventoryItem.objects.all()
    categories = InventoryCategory.objects.all()
    created_by_profiles = Profile.objects.all()

    # Initialize filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    item_id = request.GET.get('item')
    category_id = request.GET.get('category')
    created_by_id = request.GET.get('created_by')
    transaction_type = request.GET.get('transaction_type')

    # Apply filters if present
    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    if item_id:
        transactions = transactions.filter(item_id=item_id)
    if category_id:
        transactions = transactions.filter(item__category_id=category_id)
    if created_by_id:
        transactions = transactions.filter(created_by_id=created_by_id)
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)

    # Handle export to Excel
    if request.GET.get('export') == 'excel':
        return export_transactions_to_excel(transactions)

    return render(request, 'core/transaction_list.html', {
        'transactions': transactions,
        'items': items,
        'categories': categories,
        'created_by_profiles': created_by_profiles,
        'start_date': start_date,
        'end_date': end_date,
        'selected_item': item_id,
        'selected_category': category_id,
        'selected_created_by': created_by_id,
        'selected_transaction_type': transaction_type,
    })


def export_transactions_to_excel(transactions):
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Inventory Transactions'

    # Define the headers for the Excel file
    headers = ['Item', 'Quantity Changed', 'Transaction Type', 'Date', 'Created By', 'Description']
    worksheet.append(headers)

    # Write transaction data to the worksheet
    for transaction in transactions:
        worksheet.append([
            transaction.item.name,
            transaction.quantity_changed,
            transaction.transaction_type,
            transaction.date.strftime("%Y-%m-%d"),
            transaction.created_by.user.username if transaction.created_by else 'N/A',
            transaction.description if transaction.description else 'N/A',
        ])

    # Set the content type for the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'

    # Save the workbook to the response
    workbook.save(response)
    return response



@login_required(login_url='login')
def transaction_create(request):
    items = InventoryItem.objects.all()
    if request.method == 'POST':
        item_id = request.POST.get('item')
        quantity_changed = request.POST.get('quantity_changed')
        transaction_type = request.POST.get('transaction_type')
        description = request.POST.get('description')

        # Validate item and quantity
        if item_id and quantity_changed:
            transaction = InventoryTransaction.objects.create(
                item_id=item_id,
                quantity_changed=quantity_changed,
                transaction_type=transaction_type,
                description=description,
                created_by=request.user.profile  # Assuming Profile is linked to User
            )

            # Update the Inventory Item quantity
            inventory_item = InventoryItem.objects.get(pk=item_id)
            if transaction_type == 'Addition':
                inventory_item.quantity += int(quantity_changed)
            elif transaction_type == 'Removal':
                inventory_item.quantity -= int(quantity_changed)
            inventory_item.save()

            messages.success(request, "Transaction created successfully.")
            return redirect('transaction_list')
        else:
            messages.error(request, "Item and quantity are required.")

    return render(request, 'core/transaction_form.html', {'items': items})

@login_required(login_url='login')
def transaction_update(request, pk):
    transaction = get_object_or_404(InventoryTransaction, pk=pk)
    items = InventoryItem.objects.all()

    if request.method == 'POST':
        original_quantity = transaction.quantity_changed
        item_id = request.POST.get('item')
        quantity_changed = request.POST.get('quantity_changed')
        transaction_type = request.POST.get('transaction_type')
        description = request.POST.get('description')

        # Calculate the difference
        quantity_difference = int(quantity_changed) - original_quantity

        # Update the transaction fields
        transaction.item_id = item_id
        transaction.quantity_changed = quantity_changed
        transaction.transaction_type = transaction_type
        transaction.description = description
        transaction.save()

        # Update the Inventory Item quantity based on transaction type
        inventory_item = InventoryItem.objects.get(pk=item_id)
        if transaction_type == 'Addition':
            inventory_item.quantity += quantity_difference
        elif transaction_type == 'Removal':
            inventory_item.quantity -= quantity_difference

        inventory_item.save()

        messages.success(request, "Transaction updated successfully.")
        return redirect('transaction_list')

    return render(request, 'core/transaction_form.html', {'transaction': transaction, 'items': items})

@login_required(login_url='login')
def transaction_delete(request, pk):
    transaction = get_object_or_404(InventoryTransaction, pk=pk)
    inventory_item = InventoryItem.objects.get(pk=transaction.item_id)

    # Adjust inventory item quantity based on transaction type before deletion
    if transaction.transaction_type == 'Addition':
        inventory_item.quantity -= int(transaction.quantity_changed)
    elif transaction.transaction_type == 'Removal':
        inventory_item.quantity += int(transaction.quantity_changed)
    inventory_item.save()

    transaction.delete()
    messages.success(request, "Transaction deleted successfully.")
    return redirect('transaction_list')





@login_required(login_url='login')
def report_type_list(request):
    report_types = DiagnosticReportType.objects.all()
    return render(request, 'core/report_type_list.html', {'report_types': report_types})

@login_required(login_url='login')
def report_type_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')

        # Create a new DiagnosticReportType
        DiagnosticReportType.objects.create(name=name, description=description)
        messages.success(request, 'Report type created successfully.')
        return redirect('report_type_list')

    return render(request, 'core/report_type_form.html')

@login_required(login_url='login')
def report_type_update(request, pk):
    report_type = get_object_or_404(DiagnosticReportType, pk=pk)

    if request.method == 'POST':
        report_type.name = request.POST.get('name')
        report_type.description = request.POST.get('description')
        report_type.save()

        messages.success(request, 'Report type updated successfully.')
        return redirect('report_type_list')

    return render(request, 'core/report_type_form.html', {'report_type': report_type})

@login_required(login_url='login')
def report_type_delete(request, pk):
    report_type = get_object_or_404(DiagnosticReportType, pk=pk)
    report_type.delete()
    messages.success(request, 'Report type deleted successfully.')
    return redirect('report_type_list')


# DiagnosticReportParameter Views

@login_required(login_url='login')
def report_parameter_list(request):
    parameters = DiagnosticReportParameter.objects.select_related('report_type').all()
    return render(request, 'core/report_parameter_list.html', {'parameters': parameters})

@login_required(login_url='login')
def report_parameter_create(request):
    report_types = DiagnosticReportType.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        report_type_id = request.POST.get('report_type')

        # Create a new DiagnosticReportParameter
        report_type = get_object_or_404(DiagnosticReportType, pk=report_type_id)
        DiagnosticReportParameter.objects.create(name=name, description=description, report_type=report_type)
        messages.success(request, 'Report parameter created successfully.')
        return redirect('report_parameter_list')

    return render(request, 'core/report_parameter_form.html', {'report_types': report_types})

@login_required(login_url='login')
def report_parameter_update(request, pk):
    parameter = get_object_or_404(DiagnosticReportParameter, pk=pk)
    report_types = DiagnosticReportType.objects.all()

    if request.method == 'POST':
        parameter.name = request.POST.get('name')
        parameter.description = request.POST.get('description')
        parameter.report_type_id = request.POST.get('report_type')
        parameter.save()

        messages.success(request, 'Report parameter updated successfully.')
        return redirect('report_parameter_list')

    return render(request, 'core/report_parameter_form.html', {'parameter': parameter, 'report_types': report_types})

@login_required(login_url='login')
def report_parameter_delete(request, pk):
    parameter = get_object_or_404(DiagnosticReportParameter, pk=pk)
    parameter.delete()
    messages.success(request, 'Report parameter deleted successfully.')
    return redirect('report_parameter_list')




@login_required(login_url='login')
def create_diagnostic_report(request):
    if request.method == 'POST':
        # Get the basic report data
        patient_name = request.POST.get('patient_name')
        patient_age = request.POST.get('patient_age')
        gender = request.POST.get('gender')
        report_type_id = request.POST.get('report_type')
        notes = request.POST.get('notes')

        # Create the main DiagnosticReport object
        report_type = get_object_or_404(DiagnosticReportType, id=report_type_id)
        report = DiagnosticReport.objects.create(
            patient_name=patient_name,
            patient_age=patient_age,
            gender=gender,
            report_type=report_type,
            created_by=request.user.profile,  # Assuming Profile is linked to User
            notes=notes
        )

        # Save DiagnosticReportResult for each parameter
        parameter_ids = request.POST.getlist('parameter_ids[]')
        parameter_values = request.POST.getlist('parameter_values[]')
        for param_id, value in zip(parameter_ids, parameter_values):
            parameter = get_object_or_404(DiagnosticReportParameter, id=param_id)
            DiagnosticReportResult.objects.create(
                report=report,
                diagnostic_report_parameter=parameter,
                value=value
            )

        # Save DiagnosticReportImage for each uploaded image
        images = request.FILES.getlist('images[]')
        captions = request.POST.getlist('image_captions[]')
        for image, caption in zip(images, captions):
            DiagnosticReportImage.objects.create(
                report=report,
                image=image,
                caption=caption
            )

        messages.success(request, "Diagnostic report created successfully!")
        return redirect('diagnostic_report_list')  # Redirect to an appropriate view

    report_types = DiagnosticReportType.objects.all()
    parameters = DiagnosticReportParameter.objects.all()
    
    return render(request, 'core/diagnostic_report_form.html', {
        'report_types': report_types,
        'parameters': parameters,  # Pass all parameters to template
    })

@login_required(login_url='login')
def get_parameters(request, report_type_id):
    # Retrieve parameters for the given report type (for AJAX call)
    report_type = get_object_or_404(DiagnosticReportType, id=report_type_id)
    parameters = DiagnosticReportParameter.objects.filter(report_type=report_type).values('id', 'name')
    # print("parameters")
    return JsonResponse({'parameters': list(parameters)})



@login_required(login_url='login')
def diagnostic_report_list(request):
    today = timezone.now().date()
    default_start = today - timedelta(days=6)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
    except ValueError:
        start_date = default_start
    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
    except ValueError:
        end_date = today

    reports = DiagnosticReport.objects.filter(created_at__date__range=[start_date, end_date])

    return render(request, 'core/diagnostic_report_list.html', {
        'reports': reports,
        'start_date': start_date,
        'end_date': end_date,
    })

@login_required(login_url='login')
def diagnostic_report_delete(request, report_id):
    report = get_object_or_404(DiagnosticReport, id=report_id)
    report.delete()
    messages.success(request, "Diagnostic Report deleted successfully.")
    return redirect('diagnostic_report_list')



@login_required(login_url='login')
def diagnostic_report_edit(request, report_id):
    report = get_object_or_404(DiagnosticReport, id=report_id)
    
    if request.method == 'POST':
        # Update basic report details
        report.patient_name = request.POST.get('patient_name')
        report.patient_age = request.POST.get('patient_age')
        report.gender = request.POST.get('gender')
        report.notes = request.POST.get('notes')
        report_type_id = request.POST.get('report_type')
        report.report_type = DiagnosticReportType.objects.get(id=report_type_id)

        # Save updated report
        report.save()

        # Update existing parameters and values
        parameter_ids = request.POST.getlist('parameter_ids[]')
        parameter_values = request.POST.getlist('parameter_values[]')

        # Clear existing parameters for this report to avoid duplicates
        DiagnosticReportResult.objects.filter(report=report).delete()

        # Save updated parameters and values
        for param_id, value in zip(parameter_ids, parameter_values):
            parameter = DiagnosticReportParameter.objects.get(id=param_id)
            DiagnosticReportResult.objects.create(
                report=report,
                diagnostic_report_parameter=parameter,
                value=value
            )

        # Handle image updates
        if 'images[]' in request.FILES:
            images = request.FILES.getlist('images[]')
            captions = request.POST.getlist('image_captions[]')
            # Clear existing images if needed
            DiagnosticReportImage.objects.filter(report=report).delete()
            for image, caption in zip(images, captions):
                DiagnosticReportImage.objects.create(
                    report=report,
                    image=image,
                    caption=caption
                )

        messages.success(request, 'Report updated successfully!')
        return redirect('diagnostic_report_list')

    # Get all report types and parameters related to the current report type
    report_types = DiagnosticReportType.objects.all()
    parameters = DiagnosticReportParameter.objects.filter(report_type=report.report_type)
    existing_parameters = DiagnosticReportResult.objects.filter(report=report)
    images = DiagnosticReportImage.objects.filter(report=report)

    context = {
        'report': report,
        'report_types': report_types,
        'parameters': parameters,
        'existing_parameters': existing_parameters,
        'images': images,
    }

    return render(request, 'core/diagnostic_report_form.html', context)
